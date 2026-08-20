"""统一Excel生成工具 - 生成符合规范的测试用例Excel文件

负责：
- 使用固定模板生成Excel文件
- 确保15字段格式正确
- 自动输出到规范位置
- 生成规范的文件名
- 支持样式美化（表头颜色、边框、列宽）
- 支持批量导出和文件大小限制

所有Agent必须使用此工具生成Excel文件，禁止自行编写Excel生成代码。
"""

import json
import logging
import shutil
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl未安装，请运行: pip install openpyxl")

from .template_builder import ALL_FIELDS as TEMPLATE_FIELDS, _template_header_matches, get_default_template_path
from .workspace_manager import workspace_manager


class ExcelGenerator:
    """统一Excel生成器"""

    # 15字段标准顺序（单一定义源：template_builder.ALL_FIELDS）
    STANDARD_FIELDS = list(TEMPLATE_FIELDS)
    QUALITY_FIELD = "质量评分"
    # 评分、审核和重生轨迹只存在于运行时对象，不得扩展正式 Excel 表头。
    RUNTIME_FIELDS = {
        "原始评分",
        "优化后评分",
        "最终评分",
        "最终审核通过",
        "是否冷启动评分",
        "评分置信度",
        "评分历史",
        "优化次数",
        "评分门槛",
        "最终评分是否达标",
        "regeneration_count",
        "last_regenerated_at",
        "execution_count",
        "用例ID",
    }

    # 列宽配置
    COLUMN_WIDTHS = {
        "用例目录": 25,
        "用例名称": 40,
        "需求ID": 15,
        "前置条件": 40,
        "用例步骤": 50,
        "预期结果": 45,
        "用例类型": 15,
        "用例状态": 12,
        "用例等级": 10,
        "创建人": 12,
        "优先级": 10,
        "是否可自动化": 12,
        "回归测试标识": 12,
        "知识库关联": 30,
        "质量评分": 12,
    }

    @classmethod
    def _get_effective_fields(cls, extra_fields: list[str] | None = None) -> list[str]:
        """获取固定的15字段列表。

        Args:
            extra_fields: 保留的兼容参数；不允许扩展正式表头
        """
        if extra_fields:
            raise ValueError("正式Excel固定为15字段，不允许传入额外字段")
        return list(cls.STANDARD_FIELDS)

    # 模板文件路径（通过项目标记文件推断）
    TEMPLATE_PATH = get_default_template_path()

    # 文件大小限制（10MB）
    MAX_FILE_SIZE = 10 * 1024 * 1024

    @classmethod
    def generate(cls, cases: list[dict[str, Any]], output_path: str = "", requirement_name: str = "未命名需求", extra_fields: list[str] | None = None) -> str:
        """生成Excel文件（便捷方法）

        Args:
            cases: 测试用例列表
            output_path: 输出路径（可选，为空时自动生成规范路径）
            requirement_name: 需求名称（可选，默认"未命名需求"）
            extra_fields: 已废弃；正式Excel不允许额外字段

        Returns:
            str: 生成的Excel文件路径
        """
        return cls.generate_excel(test_cases=cases, output_path=output_path or None, requirement_name=requirement_name, extra_fields=extra_fields)

    @classmethod
    def execute(cls, input_data=None, **kwargs):
        """执行方法 - 支持Agent编排器调用

        Args:
            input_data: 输入数据（测试用例列表）
            **kwargs: 其他参数（requirement_name, requirement_id等）

        Returns:
            str: 生成的Excel文件路径
        """
        logger.debug(f"ExcelGenerator.execute - input_data type: {type(input_data)}")
        logger.debug(
            f"ExcelGenerator.execute - input_data: {input_data[:100] if isinstance(input_data, str) else 'Not a string'}"
        )

        if input_data is None:
            raise ValueError("input_data is required")

        # 如果输入数据不是列表，尝试解析
        if isinstance(input_data, str):
            try:
                input_data = json.loads(input_data)
            except json.JSONDecodeError as exc:
                raise ValueError("input_data 不是有效的 JSON 字符串") from exc

        # 如果是字典而不是列表，检查是否有data字段
        if isinstance(input_data, dict) and "data" in input_data:
            input_data = input_data["data"]

        # 如果仍然不是列表，抛出错误
        if not isinstance(input_data, list):
            raise ValueError(f"input_data must be a list, got {type(input_data)}")

        requirement_name = kwargs.get("requirement_name", "未命名需求")
        requirement_id = kwargs.get("requirement_id", "")
        extra_fields = kwargs.get("extra_fields")

        return cls.generate_excel(
            test_cases=input_data, requirement_name=requirement_name, requirement_id=requirement_id, extra_fields=extra_fields
        )

    @classmethod
    def generate_excel(
        cls,
        test_cases: list[dict[str, str]],
        requirement_name: str,
        requirement_id: str | None = None,
        date_str: str | None = None,
        output_path: str | None = None,
        extra_fields: list[str] | None = None,
    ) -> str:
        """生成符合规范的Excel文件

        文件名格式：需求{id}-{需求名}.xlsx
        输出路径：workspace/YYYYMMDD/

        Args:
            test_cases: 测试用例列表，每个用例是一个字典
            requirement_name: 需求名称
            requirement_id: 需求ID（可选）
            date_str: 日期字符串（可选，默认当前北京日期）
            output_path: 自定义输出路径（可选，如需自定义位置）
            extra_fields: 已废弃；正式Excel不允许额外字段

        Returns:
            str: 生成的Excel文件路径

        Raises:
            ValueError: 如果参数无效或测试用例格式错误
            FileNotFoundError: 如果模板文件不存在
        """
        cls._validate_test_cases(test_cases, extra_fields)

        if output_path is None:
            output_path = workspace_manager.generate_file_path(
                requirement_name=requirement_name,
                requirement_id=requirement_id,
                date_str=date_str,
                sub_dir=None,
            )

        # 生成Excel文件
        if Path(cls.TEMPLATE_PATH).exists():
            cls._generate_from_template(test_cases, output_path, extra_fields)
        else:
            cls._generate_new(test_cases, output_path, extra_fields)

        return str(output_path)

    @classmethod
    def _validate_test_cases(cls, test_cases: list[dict[str, str]], extra_fields: list[str] | None = None) -> None:
        """验证测试用例格式

        Args:
            test_cases: 测试用例列表
            extra_fields: 额外字段列表（可选）

        Raises:
            ValueError: 如果测试用例格式错误
        """
        if not test_cases:
            raise ValueError("测试用例列表不能为空")

        cls._get_effective_fields(extra_fields)

        for idx, case in enumerate(test_cases, start=1):
            if not isinstance(case, dict):
                raise ValueError(f"第{idx}条用例必须是对象，实际类型: {type(case).__name__}")

            unknown_fields = {
                field
                for field in set(case) - set(cls.STANDARD_FIELDS) - cls.RUNTIME_FIELDS
                if not str(field).startswith("_runtime_")
            }
            if unknown_fields:
                names = ", ".join(sorted(str(field) for field in unknown_fields))
                raise ValueError(f"第{idx}条用例包含未定义字段: {names}")

            for field in cls.STANDARD_FIELDS:
                if field not in case:
                    raise ValueError(f"第{idx}条用例缺少字段: {field}")

                value = case[field]
                if field == cls.QUALITY_FIELD:
                    if isinstance(value, bool) or not isinstance(value, (int, float)):
                        raise ValueError(f"第{idx}条用例字段'{field}'必须是0-100数字")
                    if not 0 <= value <= 100:
                        raise ValueError(f"第{idx}条用例字段'{field}'必须在0-100之间")
                elif not isinstance(value, str):
                    raise ValueError(f"第{idx}条用例字段'{field}'必须是字符串")

            # 检查用例名称
            case_name = case.get("用例名称", "").strip()
            if not case_name:
                raise ValueError(f"第{idx}条用例缺少用例名称")

    @classmethod
    def _write_header_with_style(cls, ws: Worksheet, extra_fields: list[str] | None = None) -> None:
        """写入带样式的表头

        Args:
            ws: 工作表
            extra_fields: 额外字段列表（可选）
        """
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_idx, field in enumerate(cls._get_effective_fields(extra_fields), start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 设置列宽
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = cls.COLUMN_WIDTHS.get(field, 20)

    @classmethod
    def _write_case_row_with_style(cls, ws: Worksheet, row_num: int, case: dict[str, Any], extra_fields: list[str] | None = None) -> None:
        """写入带样式的单行测试用例

        Args:
            ws: 工作表
            row_num: 行号
            case: 测试用例字典
            extra_fields: 额外字段列表（可选）
        """
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_idx, field in enumerate(cls._get_effective_fields(extra_fields), start=1):
            value = case.get(field, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            elif value is None:
                value = ""
            else:
                value = str(value)
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    @classmethod
    def _generate_from_template(cls, test_cases: list[dict[str, str]], output_path: Path, extra_fields: list[str] | None = None) -> None:
        """从模板生成Excel文件

        Args:
            test_cases: 测试用例列表
            output_path: 输出文件路径
            extra_fields: 额外字段列表（可选）
        """
        wb, ws = cls._load_or_create_workbook(output_path, cls.TEMPLATE_PATH, extra_fields)
        try:
            cls._write_test_cases_with_style(ws, test_cases, extra_fields)
            wb.save(output_path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @classmethod
    def _generate_new(cls, test_cases: list[dict[str, str]], output_path: Path, extra_fields: list[str] | None = None) -> None:
        """创建新的Excel文件（无模板时）

        Args:
            test_cases: 测试用例列表
            output_path: 输出文件路径
            extra_fields: 额外字段列表（可选）
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        cls._write_header_with_style(ws, extra_fields)

        cls._write_test_cases_with_style(ws, test_cases, extra_fields)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        wb.close()

    @classmethod
    def _write_test_cases_with_style(cls, ws: Worksheet, test_cases: list[dict[str, str]], extra_fields: list[str] | None = None) -> None:
        """将测试用例写入工作表（带样式）

        Args:
            ws: 工作表
            test_cases: 测试用例列表
            extra_fields: 额外字段列表（可选）
        """
        for row_idx, case in enumerate(test_cases, start=2):
            cls._write_case_row_with_style(ws, row_idx, case, extra_fields)

    @classmethod
    def _write_test_cases(cls, ws: Worksheet, test_cases: list[dict[str, str]]) -> None:
        """将测试用例写入工作表（兼容旧接口）

        Args:
            ws: 工作表
            test_cases: 测试用例列表
        """
        cls._write_test_cases_with_style(ws, test_cases)

    @classmethod
    def validate_excel(cls, file_path: str) -> tuple[bool, str]:
        """验证Excel文件是否符合规范

        Args:
            file_path: Excel文件路径

        Returns:
            tuple[bool, str]: (是否符合规范, 错误信息)
        """
        p = Path(file_path)
        if not p.exists():
            return False, f"文件不存在: {file_path}"
        if p.suffix.lower() != ".xlsx":
            return False, f"文件格式必须为.xlsx: {file_path}"

        wb = None
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb["测试用例"]

            width = max(ws.max_column, len(cls.STANDARD_FIELDS))
            header = [ws.cell(row=1, column=c).value for c in range(1, width + 1)]
            while header and header[-1] in (None, ""):
                header.pop()

            for idx, (expected, actual) in enumerate(zip(cls.STANDARD_FIELDS, header)):
                if expected != actual:
                    return False, f"第{idx + 1}列表头错误: 期望'{expected}', 实际'{actual}'"
            if header != cls.STANDARD_FIELDS:
                return False, f"Excel表头必须严格为15字段，实际字段数: {len(header)}"

            row_count = ws.max_row - 1
            if row_count == 0:
                return False, "Excel文件没有数据行"

            return True, ""

        except Exception as e:
            return False, f"验证Excel文件失败: {str(e)}"
        finally:
            if wb is not None:
                wb.close()

    @classmethod
    def create_empty_case(
        cls,
        case_directory: str = "",
        case_name: str = "",
        requirement_id: str = "",
        precondition: str = "",
        case_steps: str = "",
        expected_result: str = "",
        case_type: str = "功能测试",
        case_status: str = "正常",
        case_level: str = "中",
        creator: str = "",
        priority: str = "P2",
        is_automation: str = "否",
        regression_flag: str = "否",
        knowledge_link: str = "",
    ) -> dict[str, str]:
        """创建空的测试用例字典

        Args:
            case_directory: 用例目录
            case_name: 用例名称
            requirement_id: 需求ID
            precondition: 前置条件
            case_steps: 用例步骤
            expected_result: 预期结果
            case_type: 用例类型
            case_status: 用例状态
            case_level: 用例等级
            creator: 创建人
            priority: 优先级
            is_automation: 是否可自动化
            regression_flag: 回归测试标识
            knowledge_link: 知识库关联

        Returns:
            dict[str, str]: 测试用例字典
        """
        return {
            "用例目录": case_directory,
            "用例名称": case_name,
            "需求ID": requirement_id,
            "前置条件": precondition,
            "用例步骤": case_steps,
            "预期结果": expected_result,
            "用例类型": case_type,
            "用例状态": case_status,
            "用例等级": case_level,
            "创建人": creator,
            "优先级": priority,
            "是否可自动化": is_automation,
            "回归测试标识": regression_flag,
            "知识库关联": knowledge_link,
            "质量评分": 0.0,
        }

    @staticmethod
    def generate_case_id() -> str:
        """生成唯一用例编号

        Returns:
            str: 用例编号，格式为 TC-xxxxxxxx
        """
        return f"TC-{uuid.uuid4().hex[:8].upper()}"

    @classmethod
    def _load_or_create_workbook(cls, output_path: Path | str, template_path: str | None = None, extra_fields: list[str] | None = None) -> tuple[Any, Any]:
        """加载模板或创建工作簿，返回 (wb, ws)

        Args:
            output_path: 输出文件路径
            template_path: 模板文件路径（可选）
            extra_fields: 额外字段列表（可选）

        Returns:
            tuple[Any, Any]: (workbook, worksheet)
        """
        output_path = Path(output_path)
        if template_path and Path(template_path).exists():
            if not _template_header_matches(template_path):
                raise ValueError(f"模板表头不符合15字段契约: {template_path}")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(template_path, output_path)
            wb = load_workbook(output_path)
            ws = wb["测试用例"]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            cls._write_header_with_style(ws, extra_fields)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "测试用例"
            cls._write_header_with_style(ws, extra_fields)
        return wb, ws

    @classmethod
    def export_cases(
        cls, cases: list[dict[str, Any]], output_path: str, template_path: str = None, extra_fields: list[str] | None = None
    ) -> dict[str, Any]:
        """导出测试用例到Excel（支持文件大小限制）

        Args:
            cases: 测试用例列表
            output_path: 输出文件路径
            template_path: 模板文件路径（可选）
            extra_fields: 额外字段列表（可选）

        Returns:
            导出结果字典
        """
        if not cases:
            return {"success": False, "error": "无测试用例可导出"}

        try:
            cls._validate_test_cases(cases, extra_fields)

            output_path = Path(output_path)
            wb, ws = cls._load_or_create_workbook(output_path, template_path, extra_fields)

            try:
                for row_idx, case in enumerate(cases, start=2):
                    cls._write_case_row_with_style(ws, row_idx, case, extra_fields)

                output_path.parent.mkdir(parents=True, exist_ok=True)
                wb.save(str(output_path))

                file_size = output_path.stat().st_size
                if file_size > cls.MAX_FILE_SIZE:
                    output_path.unlink(missing_ok=True)
                    return {"success": False, "error": f"导出文件超过{cls.MAX_FILE_SIZE // (1024 * 1024)}MB限制"}

                return {
                    "success": True,
                    "path": str(output_path),
                    "case_count": len(cases),
                    "file_size": file_size,
                    "export_time": datetime.now().isoformat(),
                }
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

        except Exception as e:
            logger.exception("导出测试用例失败")
            return {"success": False, "error": str(e)}

    @classmethod
    def export_single_case(cls, case: dict[str, Any], output_path: str, extra_fields: list[str] | None = None) -> dict[str, Any]:
        """单例导出模式

        Args:
            case: 测试用例字典
            output_path: 输出文件路径
            extra_fields: 额外字段列表（可选）

        Returns:
            导出结果字典
        """
        return cls.export_cases([case], output_path, extra_fields=extra_fields)

    @classmethod
    def batch_export(
        cls, cases: list[dict[str, Any]], output_dir: str, batch_size: int = 100, requirement_name: str = "未命名需求", extra_fields: list[str] | None = None
    ) -> list[dict[str, Any]]:
        """批量导出模式

        Args:
            cases: 测试用例列表
            output_dir: 输出目录
            batch_size: 每批数量，默认100
            requirement_name: 需求名称
            extra_fields: 额外字段列表（可选）

        Returns:
            各批次导出结果列表
        """
        results = []
        total_cases = len(cases)
        batch_count = (total_cases + batch_size - 1) // batch_size
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir_path = Path(output_dir)

        for batch_idx in range(batch_count):
            start = batch_idx * batch_size
            end = min((batch_idx + 1) * batch_size, total_cases)
            batch_cases = cases[start:end]

            filename = f"{requirement_name}_batch_{batch_idx + 1}_{timestamp}.xlsx"
            output_path = str(output_dir_path / filename)

            result = cls.export_cases(batch_cases, output_path, extra_fields=extra_fields)
            result["batch_number"] = batch_idx + 1
            result["total_batches"] = batch_count
            result["cases_in_batch"] = len(batch_cases)
            results.append(result)

        return results

    @staticmethod
    def read_excel_worksheet(file_path: str) -> list[dict[str, str]]:
        """读取工作区Excel文件中的测试用例

        Args:
            file_path: Excel文件路径

        Returns:
            测试用例列表，每个元素包含 case_name, steps, expected_result
        """
        from openpyxl import load_workbook

        wb = None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            if ws.max_row < 2:
                return []

            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            try:
                steps_col = header.index("用例步骤")
                expected_col = header.index("预期结果")
                name_col = header.index("用例名称")
            except ValueError:
                return []

            results = []
            for row in range(2, ws.max_row + 1):
                case_name = str(ws.cell(row=row, column=name_col + 1).value or "")
                steps = str(ws.cell(row=row, column=steps_col + 1).value or "")
                expected = str(ws.cell(row=row, column=expected_col + 1).value or "")
                results.append({
                    "case_name": case_name,
                    "steps": steps.strip(),
                    "expected_result": expected.strip(),
                })
            return results
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass


# 全局实例
excel_generator = ExcelGenerator()
