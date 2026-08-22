"""固定测试用例模板与 15 字段 schema 契约。

``fixtures/templates/测试用例模板.xlsx`` 是唯一正式模板来源。运行时只读校验
该文件，绝不在 ``fixtures`` 下创建、覆盖或修复文件。
"""

from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path
from typing import Any

IMPORT_FIELDS: list[str] = [
    "用例目录", "用例名称", "需求ID", "前置条件", "用例步骤", "预期结果",
    "用例类型", "用例状态", "用例等级", "创建人", "优先级",
]
EXTENSION_FIELDS: list[str] = ["是否可自动化", "回归测试标识", "知识库关联", "质量评分"]

# 正式表头的唯一 Python 定义源；ExcelGenerator 只引用它，不再维护副本。
ALL_FIELDS: list[str] = IMPORT_FIELDS + EXTENSION_FIELDS
WORKSHEET_NAME = "测试用例"
TEMPLATE_FILENAME = "测试用例模板.xlsx"
SCHEMA_VERSION = "1.0"
QUALITY_SCORE_MIN = 0.0
QUALITY_SCORE_MAX = 100.0

RUNTIME_QUALITY_FIELDS = {
    "original_score", "optimized_score", "final_score", "score_history",
    "score_threshold", "is_cold_start", "confidence", "optimization_attempts",
    "needs_human_review", "final_audit_passed",
}
RUNTIME_COVERAGE_FIELDS = {
    "business_rules", "business_objects", "normal_scenarios", "abnormal_scenarios",
    "boundary_scenarios", "rollback_scenarios", "exclusions",
}
RUNTIME_REGENERATION_FIELDS = {"count", "last_regenerated_at"}
RUNTIME_FIELDS = {
    "_runtime_quality", "_runtime_quality_version", "_runtime_coverage_matrix",
    "_runtime_coverage_matrix_version", "_runtime_regeneration",
}

# 旧实现曾把这些字段直接写进用例字典；正式 schema 明确拒绝它们。
LEGACY_RUNTIME_FIELDS = {
    "原始评分", "优化后评分", "最终评分", "最终审核通过", "是否冷启动评分",
    "评分置信度", "评分历史", "优化次数", "评分门槛", "最终评分是否达标",
    "needs_human_review", "regeneration_count", "last_regenerated_at",
    "execution_count", "用例ID",
}

COLUMN_WIDTHS = {
    "用例目录": 25, "用例名称": 40, "需求ID": 12, "前置条件": 30,
    "用例步骤": 40, "预期结果": 35, "用例类型": 12, "用例状态": 10,
    "用例等级": 10, "创建人": 12, "优先级": 8, "是否可自动化": 12,
    "回归测试标识": 15, "知识库关联": 15, "质量评分": 12,
}


def get_default_template_path() -> str:
    """返回仓库内唯一正式模板的绝对路径。"""

    project_root = Path(__file__).resolve().parents[3]
    return str(project_root / "fixtures" / "templates" / TEMPLATE_FILENAME)


def _template_header_matches(template_path: str) -> bool:
    """严格校验模板工作表、工作表数量和 15 列表头。"""

    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    workbook = None
    try:
        workbook = load_workbook(template_path, read_only=True, data_only=False)
        if workbook.sheetnames != [WORKSHEET_NAME]:
            return False
        worksheet = workbook[WORKSHEET_NAME]
        if worksheet.max_column != len(ALL_FIELDS):
            return False
        header = [worksheet.cell(row=1, column=index).value for index in range(1, len(ALL_FIELDS) + 1)]
        return header == ALL_FIELDS
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.close()


def ensure_template(template_path: str | None = None) -> str:
    """只读确认模板存在且符合固定 schema。

    ``template_path`` 仅保留为兼容参数；传入非默认路径会被拒绝。本函数不会
    创建或修改任何文件。
    """

    canonical = Path(get_default_template_path()).resolve()
    path = Path(template_path or canonical).resolve()
    if path != canonical:
        raise ValueError("正式模板只有唯一来源: fixtures/templates/测试用例模板.xlsx")
    if not path.exists():
        raise FileNotFoundError(f"固定测试用例模板不存在，禁止运行时创建: {path}")
    if not path.is_file() or not _template_header_matches(str(path)):
        raise ValueError(f"模板必须严格为单工作表15字段契约: {path}")
    return str(path)


def _validate_score(value: Any, field_name: str) -> None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"字段'{field_name}'必须是0-100数字")
    if not QUALITY_SCORE_MIN <= float(value) <= QUALITY_SCORE_MAX:
        raise ValueError(f"字段'{field_name}'必须在0-100之间")


def _validate_runtime_quality(value: Any) -> None:
    if not isinstance(value, Mapping):
        raise ValueError("_runtime_quality必须是对象")
    unknown = set(value) - RUNTIME_QUALITY_FIELDS
    if unknown:
        raise ValueError(f"_runtime_quality包含未定义字段: {', '.join(sorted(map(str, unknown)))}")
    for name in ("original_score", "optimized_score", "final_score"):
        if name in value and value[name] is not None:
            _validate_score(value[name], name)
    if "score_threshold" in value and float(value["score_threshold"]) != 85.0:
        raise ValueError("_runtime_quality.score_threshold必须为85")
    if "confidence" in value and not 0 <= float(value["confidence"]) <= 1:
        raise ValueError("_runtime_quality.confidence必须在0-1之间")
    if "score_history" in value and not isinstance(value["score_history"], list):
        raise ValueError("_runtime_quality.score_history必须是数组")
    if "optimization_attempts" in value and (
        isinstance(value["optimization_attempts"], bool) or int(value["optimization_attempts"]) < 0
    ):
        raise ValueError("_runtime_quality.optimization_attempts必须是非负整数")
    for name in ("is_cold_start", "needs_human_review", "final_audit_passed"):
        if name in value and not isinstance(value[name], bool):
            raise ValueError(f"_runtime_quality.{name}必须是布尔值")


def _validate_runtime_coverage(value: Any) -> None:
    if not isinstance(value, Mapping):
        raise ValueError("_runtime_coverage_matrix必须是对象")
    unknown = set(value) - RUNTIME_COVERAGE_FIELDS
    if unknown:
        raise ValueError(f"_runtime_coverage_matrix包含未定义字段: {', '.join(sorted(map(str, unknown)))}")
    for name, items in value.items():
        if not isinstance(items, list) or any(not isinstance(item, str) for item in items):
            raise ValueError(f"_runtime_coverage_matrix.{name}必须是字符串数组")


def _validate_runtime_regeneration(value: Any) -> None:
    if not isinstance(value, Mapping):
        raise ValueError("_runtime_regeneration必须是对象")
    unknown = set(value) - RUNTIME_REGENERATION_FIELDS
    if unknown:
        raise ValueError(f"_runtime_regeneration包含未定义字段: {', '.join(sorted(map(str, unknown)))}")
    if "count" in value and (isinstance(value["count"], bool) or int(value["count"]) < 0):
        raise ValueError("_runtime_regeneration.count必须是非负整数")
    if "last_regenerated_at" in value and not isinstance(value["last_regenerated_at"], str):
        raise ValueError("_runtime_regeneration.last_regenerated_at必须是字符串")


def validate_case_fields(case: Mapping[str, Any], item_label: str = "用例") -> None:
    """校验正式 15 字段及受控运行时命名空间。"""

    if not isinstance(case, Mapping):
        raise ValueError(f"{item_label}必须是对象")
    unknown = set(case) - set(ALL_FIELDS) - RUNTIME_FIELDS
    if unknown:
        names = ", ".join(sorted(map(str, unknown)))
        raise ValueError(f"{item_label}包含未定义字段: {names}")
    missing = [field for field in ALL_FIELDS if field not in case]
    if missing:
        raise ValueError(f"{item_label}缺少字段: {', '.join(missing)}")

    for field in ALL_FIELDS:
        value = case[field]
        if field == "质量评分":
            _validate_score(value, field)
        elif not isinstance(value, str):
            raise ValueError(f"{item_label}字段'{field}'必须是字符串")
    if not case["用例名称"].strip():
        raise ValueError(f"{item_label}缺少用例名称")

    if "_runtime_quality" in case:
        _validate_runtime_quality(case["_runtime_quality"])
    if "_runtime_quality_version" in case and case["_runtime_quality_version"] != SCHEMA_VERSION:
        raise ValueError("_runtime_quality_version版本不匹配")
    if "_runtime_coverage_matrix" in case:
        _validate_runtime_coverage(case["_runtime_coverage_matrix"])
    if "_runtime_coverage_matrix_version" in case and case["_runtime_coverage_matrix_version"] != SCHEMA_VERSION:
        raise ValueError("_runtime_coverage_matrix_version版本不匹配")
    if "_runtime_regeneration" in case:
        _validate_runtime_regeneration(case["_runtime_regeneration"])
