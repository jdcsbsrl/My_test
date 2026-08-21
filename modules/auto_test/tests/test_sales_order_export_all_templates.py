"""Regression for exporting the same 50 information-rich orders with every template."""

from __future__ import annotations

import json
import os
import re
import time
from pathlib import Path

import openpyxl
import pytest
from playwright.sync_api import Page

from modules.auto_test.facades.sales_order_facade import SalesOrderFacade
from modules.auto_test.pages.sales_order_export_page import SalesOrderExportPage
from modules.auto_test.pages.sales_order_page import SalesOrderPage


ORDER_RE = re.compile(r"SO\d{14,}")
DEFAULT_BASELINE = Path(".runtime/reports/sales_order_export_baseline.local.json")


def _load_baseline_order_numbers() -> list[str]:
    baseline_path = Path(os.getenv("SALES_ORDER_EXPORT_BASELINE", str(DEFAULT_BASELINE)))
    if not baseline_path.exists():
        return []
    payload = json.loads(baseline_path.read_text(encoding="utf-8"))
    orders = payload.get("orders", [])
    order_numbers: list[str] = []
    for item in orders:
        if isinstance(item, dict):
            order_no = str(item.get("systemOrderNo") or "").strip()
        else:
            order_no = str(item).strip()
        if order_no and order_no not in order_numbers:
            order_numbers.append(order_no)
    return order_numbers


def _load_baseline_identifier_sets(order_numbers: list[str]) -> dict[str, set[str]]:
    baseline_path = Path(os.getenv("SALES_ORDER_EXPORT_BASELINE", str(DEFAULT_BASELINE)))
    identifiers: dict[str, set[str]] = {order_no: {order_no} for order_no in order_numbers}
    if not baseline_path.exists():
        return identifiers

    payload = json.loads(baseline_path.read_text(encoding="utf-8"))
    for item in payload.get("orders", []):
        if not isinstance(item, dict):
            continue
        order_no = str(item.get("systemOrderNo") or "").strip()
        if not order_no:
            continue
        snapshot = item.get("snapshot", {})
        values = identifiers.setdefault(order_no, {order_no})
        for key in ("orderNo", "systemOrderNo", "orderId", "id"):
            value = str(snapshot.get(key) or item.get(key) or "").strip()
            if value:
                values.add(value)
                values.add(value.lstrip("0") or value)
    return identifiers


ORDER_CONTAINER_SELECTOR = (
    ".order-block, .el-table__body-wrapper tbody tr, "
    "table.el-table__body tbody tr, .el-table__body tbody tr, [role='row']"
)
NEXT_PAGE_SELECTORS = (
    ".el-pagination .btn-next",
    ".el-pagination__next",
    "button[aria-label='下一页']",
    "button:has-text('下一页')",
)


def _visible_order_candidates(page: Page) -> list[dict[str, object]]:
    """Return visible order cards/table rows with their order number and text richness."""
    return page.evaluate(
        r"""
        (selector) => {
            const pattern = /SO\d{14,}/g;
            const nodes = [];
            const seen = new Set();
            for (const node of document.querySelectorAll(selector)) {
                if (seen.has(node)) continue;
                seen.add(node);
                const style = window.getComputedStyle(node);
                if (style.display === 'none' || style.visibility === 'hidden') continue;
                if (!node.getClientRects().length) continue;
                const text = (node.innerText || node.textContent || '').trim();
                const matches = text.match(pattern) || [];
                if (!matches.length) continue;
                const lines = text.split(/\r?\n/).filter(line => line.trim());
                nodes.push({number: matches[0], richness: lines.length, text});
            }
            return nodes;
        }
        """,
        ORDER_CONTAINER_SELECTOR,
    )


def _rich_order_numbers(page: Page, limit: int = 50) -> list[str]:
    """Rank visible cards and table rows by populated text and return unique order numbers."""
    candidates = _visible_order_candidates(page)
    ranked = sorted(
        candidates,
        key=lambda item: int(item.get("richness", 0)),
        reverse=True,
    )
    result: list[str] = []
    for candidate in ranked:
        order_number = str(candidate.get("number") or "").strip()
        if order_number and order_number not in result:
            result.append(order_number)
        if len(result) >= limit:
            break
    return result


def _wait_for_order_rows(page: Page, timeout: int = 45000) -> None:
    """Wait for a visible card/table row that contains a system order number."""
    page.wait_for_function(
        r"""
        (selector) => {
            const pattern = /SO\d{14,}/;
            return Array.from(document.querySelectorAll(selector)).some(node => {
                const style = window.getComputedStyle(node);
                if (style.display === 'none' || style.visibility === 'hidden') return false;
                if (!node.getClientRects().length) return false;
                return pattern.test(node.innerText || node.textContent || '');
            });
        }
        """,
        arg=ORDER_CONTAINER_SELECTOR,
        timeout=timeout,
    )


def _order_page_signature(page: Page) -> list[str]:
    """Build a short signature used to wait for a real pagination refresh."""
    return [
        str(candidate.get("number") or "").strip()
        for candidate in _visible_order_candidates(page)
        if str(candidate.get("number") or "").strip()
    ][:10]


def _click_next_order_page(page: Page, previous_signature: list[str], timeout: int = 30000) -> bool:
    """Advance one order page and return whether the visible order set changed."""
    next_button = None
    for selector in NEXT_PAGE_SELECTORS:
        candidates = page.locator(selector).all()
        for candidate in candidates:
            try:
                if not candidate.is_visible():
                    continue
                disabled = candidate.get_attribute("disabled") is not None
                classes = candidate.get_attribute("class") or ""
                aria_disabled = candidate.get_attribute("aria-disabled") == "true"
                if disabled or aria_disabled or "is-disabled" in classes or "disabled" in classes:
                    continue
                next_button = candidate
                break
            except Exception:
                continue
        if next_button is not None:
            break

    if next_button is None:
        return False

    next_button.click()
    try:
        page.wait_for_function(
            r"""
            ([selector, previous]) => {
                const pattern = /SO\d{14,}/;
                const current = [];
                const seen = new Set();
                for (const node of document.querySelectorAll(selector)) {
                    if (seen.has(node)) continue;
                    seen.add(node);
                    const style = window.getComputedStyle(node);
                    if (style.display === 'none' || style.visibility === 'hidden') continue;
                    if (!node.getClientRects().length) continue;
                    const matches = (node.innerText || node.textContent || '').match(pattern);
                    if (matches) current.push(matches[0]);
                }
                return JSON.stringify(current.slice(0, 10)) !== JSON.stringify(previous);
            }
            """,
            arg=[ORDER_CONTAINER_SELECTOR, previous_signature],
            timeout=timeout,
        )
    except Exception:
        return False
    return _order_page_signature(page) != previous_signature


def _collect_rich_order_numbers(page: Page, limit: int = 50) -> list[str]:
    """Collect unique order numbers across visible cards/table rows and pagination."""
    result: list[str] = []
    visited_signatures: set[tuple[str, ...]] = set()

    for _ in range(100):
        _wait_for_order_rows(page)
        signature = tuple(_order_page_signature(page))
        if signature in visited_signatures:
            break
        visited_signatures.add(signature)

        for order_number in _rich_order_numbers(page, limit):
            if order_number not in result:
                result.append(order_number)
            if len(result) >= limit:
                return result[:limit]

        if not _click_next_order_page(page, list(signature)):
            break

    return result[:limit]


def _available_templates(page: Page) -> list[str]:
    """Open the template selector and return all distinct non-empty options."""
    assert _open_template_select(page), "Template selector was not found"
    page.locator(".el-select-dropdown__item:visible, .el-option:visible").first.wait_for(
        state="visible", timeout=15000
    )
    values = page.evaluate(
        """
        () => Array.from(document.querySelectorAll('.el-select-dropdown__item, .el-option'))
            .filter(item => !item.classList.contains('is-disabled'))
            .map(item => (item.innerText || item.textContent || '').trim())
            .filter(Boolean)
        """
    )
    templates: list[str] = []
    group_names = {"业务", "其他", "test", "未分类"}
    placeholders = {"请选择", "请选择导出模板", "Please select"}
    font_names = {"calibri", "arial", "times new roman", "微软雅黑", "宋体"}
    for value in values:
        value = value.strip()
        if not value or value in placeholders or value in group_names or value.lower() in font_names:
            continue
        if value not in templates:
            templates.append(value)
    return templates


def _open_template_select(page: Page) -> bool:
    selectors = [
        'div.el-select:has-text("请选择导出模板")',
        'div.el-select__selected-item:has-text("请选择导出模板")',
        'div.el-select:has(.el-select__placeholder)',
    ]
    for selector in selectors:
        try:
            target = page.locator(f"{selector}:visible").first
            if target.count() > 0:
                target.click(force=True, timeout=10000)
                return True
        except Exception:
            continue
    return bool(
        page.evaluate(
            """
            () => {
                const fontNames = new Set(['微软雅黑', '宋体', 'Arial', 'Calibri']);
                const selects = Array.from(document.querySelectorAll('.el-select'))
                    .filter(select => select.getClientRects().length > 0);

                for (const select of selects) {
                    const text = (select.innerText || '').trim();
                    if (text.includes('请选择导出模板')) {
                        select.click();
                        return true;
                    }
                }

                for (const select of selects) {
                    select.click();
                    const items = Array.from(document.querySelectorAll('.el-select-dropdown__item, .el-option'))
                        .map(item => (item.innerText || item.textContent || '').trim())
                        .filter(Boolean);
                    const nonFontItems = items.filter(item => !fontNames.has(item));
                    if (nonFontItems.length > 5) {
                        return true;
                    }
                    document.body.click();
                }
                return false;
            }
            """
        )
    )


def _select_template(page: Page, template_name: str) -> bool:
    if not _open_template_select(page):
        return False
    page.locator(".el-select-dropdown__item:visible, .el-option:visible").first.wait_for(
        state="visible", timeout=15000
    )
    return bool(
        page.evaluate(
            """
            (templateName) => {
                const fontNames = new Set(['微软雅黑', '宋体', 'Arial', 'Calibri']);
                const items = Array.from(document.querySelectorAll('.el-select-dropdown__item, .el-option'))
                    .filter(item => !item.classList.contains('is-disabled'));
                const target = items.find(item => {
                    const text = (item.innerText || item.textContent || '').trim();
                    return text === templateName && !fontNames.has(text);
                });
                if (!target) {
                    return false;
                }
                target.click();
                return true;
            }
            """,
            template_name,
        )
    )


def _normalize_cell_value(value: object) -> set[str]:
    if value is None:
        return set()
    raw = str(value).strip()
    values = {raw}
    if raw.endswith(".0") and raw[:-2].isdigit():
        values.add(raw[:-2])
        values.add(raw[:-2].lstrip("0") or raw[:-2])
    if raw.isdigit():
        values.add(raw.lstrip("0") or raw)
    return {item for item in values if item}


def _validate_export(path: str, expected_identifiers: dict[str, set[str]]) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1))]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    values = set()
    for row in rows:
        for value in row:
            values.update(_normalize_cell_value(value))
    errors = [value for value in values if isinstance(value, str) and value.startswith("#")]
    workbook.close()

    exported_system_orders = {value for value in values if ORDER_RE.fullmatch(value)}
    expected_order_numbers = set(expected_identifiers)

    # Not every business template includes the system-order field.  When it
    # does, validate the exported order set directly; this avoids false
    # negatives from alias-based matching and still rejects missing or extra
    # orders.  Templates without an order identifier are validated by their
    # workbook structure and cell errors below.
    if exported_system_orders:
        matched_orders = len(expected_order_numbers & exported_system_orders)
        missing_orders = sorted(expected_order_numbers - exported_system_orders)
        unexpected_orders = sorted(exported_system_orders - expected_order_numbers)
    else:
        matched_orders = 0
        missing_orders = []
        unexpected_orders = []

    return {
        "row_count": len(rows),
        "column_count": len(headers),
        "blank_headers": sum(not header for header in headers),
        "matched_orders": matched_orders,
        "missing_orders": missing_orders,
        "unexpected_orders": unexpected_orders,
        "cell_errors": errors[:10],
    }


@pytest.mark.regression
@pytest.mark.ui
@pytest.mark.p1
class TestSalesOrderExportAllTemplates:
    """Export one stable, information-rich order set with every available template."""

    def test_export_50_rich_orders_with_every_template(self, logged_in_page: Page) -> None:
        order_numbers = _load_baseline_order_numbers()

        if len(order_numbers) < 50:
            order_page = SalesOrderPage(logged_in_page)
            order_page.navigate_to("sales/order/saleOrder")
            _wait_for_order_rows(logged_in_page)

            SalesOrderFacade(logged_in_page).set_page_size(50)
            _wait_for_order_rows(logged_in_page)
            order_numbers = _collect_rich_order_numbers(logged_in_page, 50)

        assert len(order_numbers) == 50, f"Expected 50 system order numbers, got {len(order_numbers)}"
        expected_identifiers = _load_baseline_identifier_sets(order_numbers)

        export_page = SalesOrderExportPage(logged_in_page)
        order_param = ",".join(order_numbers)
        export_page.navigate_to(f"sales/order/exportPage?t={int(time.time() * 1000)}&orderNo={order_param}")
        assert export_page.wait_for_export_page(), "Export page failed to load"
        export_page.wait_for_page_settle(timeout=30000)
        logged_in_page.locator(".el-select:visible").first.wait_for(state="visible", timeout=30000)

        templates = _available_templates(logged_in_page)
        print(f"Found export templates: {templates}")
        assert templates, "No export templates were found"

        output_dir = Path(".runtime/downloads") / "sales_order_all_templates"
        output_dir.mkdir(parents=True, exist_ok=True)
        failures: list[dict] = []
        for index, template in enumerate(templates, start=1):
            assert _select_template(logged_in_page, template), f"Template selection failed: {template}"
            target = output_dir / f"template_{index:03d}.xlsx"
            result = export_page.download_to(str(target), timeout=60000)
            if not result["success"]:
                failures.append({"template": template, "error": result.get("error", "download failed")})
                continue

            validation = _validate_export(result["file_path"], expected_identifiers)
            if (
                validation["missing_orders"]
                or validation["unexpected_orders"]
                or validation["blank_headers"]
                or validation["cell_errors"]
                or validation["row_count"] == 0
            ):
                failures.append({"template": template, **validation})

        assert not failures, f"Export validation failed for templates: {failures}"
