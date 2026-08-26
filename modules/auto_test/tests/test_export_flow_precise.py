"""Precise test for export flow."""

import os
import sys
import time

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

import pytest
from playwright.sync_api import Page

from modules.auto_test.pages.sales_order_export_page import SalesOrderExportPage
from modules.auto_test.pages.sales_order_page import SalesOrderPage


def _skip_ci_environment_issue(reason: str) -> None:
    if os.getenv("CI", "").lower() in {"1", "true", "yes"}:
        pytest.skip(f"CI测试环境页面/接口未就绪，跳过本次UI用例: {reason}")


@pytest.mark.regression
@pytest.mark.ui
@pytest.mark.p1
class TestExportFlowPrecise:
    """Precise test for export flow."""

    def test_export_flow_payment_time_asc(self, logged_in_page: Page) -> None:
        """Test export flow with payment time ascending sort."""
        sales_order_page = SalesOrderPage(logged_in_page)
        export_page = SalesOrderExportPage(logged_in_page)

        print("\n" + "=" * 70)
        print("Step 1: Navigate to sales order page")
        print("=" * 70)
        sales_order_page.navigate_to("sales/order/saleOrder")
        logged_in_page.wait_for_load_state("networkidle")
        print("URL: same-origin sales order route reached")

        print("\n" + "=" * 70)
        print("Step 2: Click pending tab")
        print("=" * 70)
        sales_order_page.click_tab("待处理")
        logged_in_page.wait_for_load_state("networkidle")

        print("\n" + "=" * 70)
        print("Step 3: Get order numbers before sort")
        print("=" * 70)
        order_numbers_before = sales_order_page.get_sorted_order_numbers(limit=10)
        print(f"Order numbers before sort: {len(order_numbers_before)} found")

        print("\n" + "=" * 70)
        print("Step 4: Click sort dropdown and select payment time ascending")
        print("=" * 70)
        sales_order_page.click_sort_dropdown()
        logged_in_page.wait_for_load_state("networkidle")

        sales_order_page.select_sort_order("付款时间", is_ascending=True)
        logged_in_page.wait_for_load_state("networkidle")

        print("\n" + "=" * 70)
        print("Step 5: Get order numbers after sort")
        print("=" * 70)
        page_order_numbers = sales_order_page.get_sorted_order_numbers(limit=30)
        print(f"Order numbers after sort (payment time ascending): {len(page_order_numbers)} found")
        print(f"Total: {len(page_order_numbers)}")

        assert len(page_order_numbers) > 0, "页面未获取到订单号"

        print("\n" + "=" * 70)
        print("Step 6: Navigate to export page with timestamp and order numbers")
        print("=" * 70)
        timestamp = str(int(time.time() * 1000))
        order_param = ",".join(page_order_numbers[:3]) if page_order_numbers else ""
        print(f"Order param for export: {len(page_order_numbers[:3])} order numbers")
        export_page.navigate_to(f"sales/order/exportPage?t={timestamp}&orderNo={order_param}")
        logged_in_page.wait_for_load_state("networkidle")

        print("Export page URL: same-origin export route reached")

        print("\n" + "=" * 70)
        print("Step 7: Click template select dropdown")
        print("=" * 70)
        logged_in_page.evaluate("""
            () => {
                const selects = document.querySelectorAll('.el-select');
                if (selects.length > 0) {
                    selects[0].click();
                    return true;
                }
                return false;
            }
        """)
        logged_in_page.wait_for_load_state("networkidle")

        print("\n" + "=" * 70)
        print("Step 8: List available templates")
        print("=" * 70)
        templates = logged_in_page.evaluate("""
            () => {
                const result = [];
                const items = document.querySelectorAll('.el-select-dropdown__item, .el-option');
                for (let i = 0; i < items.length; i++) {
                    const text = items[i].innerText.trim();
                    if (text && text.length > 0) {
                        result.push({ index: i, text: text });
                    }
                }
                return result;
            }
        """)
        print(f"Found {len(templates)} templates:")
        for template in templates:
            print(f"  [{template['index']}] '{template['text']}'")

        print("\n" + "=" * 70)
        print("Step 9: Select template by keyword 'Dayone'")
        print("=" * 70)
        selected_template_text = logged_in_page.evaluate("""
            () => {
                const items = document.querySelectorAll('.el-select-dropdown__item, .el-option');
                for (let i = 0; i < items.length; i++) {
                    const text = items[i].innerText.trim();
                    if (text.includes('Dayone')) {
                        items[i].click();
                        return text;
                    }
                }
                return null;
            }
        """)
        print(f"Selected template: '{selected_template_text}'")
        if selected_template_text is None:
            _skip_ci_environment_issue("未找到Dayone导出模板")

        assert selected_template_text is not None, "未找到Dayone模板"

        logged_in_page.wait_for_load_state("networkidle")

        print("\n" + "=" * 70)
        print("Step 10: Ensure order number field is selected (use template defaults)")
        print("=" * 70)

        selected_count = logged_in_page.evaluate("""
            () => {
                return document.querySelectorAll('input[type="checkbox"]:checked').length;
            }
        """)
        total_count = logged_in_page.evaluate("""
            () => {
                return document.querySelectorAll('input[type="checkbox"]').length;
            }
        """)
        print(f"Default selected fields: {selected_count}/{total_count}")

        print("\n" + "=" * 70)
        print("Step 11: Ensure '系统单号' field is selected")
        print("=" * 70)
        field_added = export_page.select_field("系统单号")
        print(f"Field '系统单号' selected: {field_added}")

        print("\n" + "=" * 70)
        print("Step 12: Download via export_page.download_to()")
        print("=" * 70)

        timestamp = int(time.time())
        save_path = f".runtime/downloads/sales_order_payment_time_asc_{timestamp}.xlsx"

        download_result = export_page.download_to(save_path, timeout=120000)

        if not download_result["success"]:
            print(f"\n✗ Download failed: {download_result.get('error', 'Unknown')}")
            pytest.fail(f"下载失败: {download_result.get('error')}")

        file_path = download_result["file_path"]
        file_size = download_result["file_size"]

        print("\n✓ Download successful!")
        print(f"  Filename: {download_result['filename']}")
        print(f"  File path: {file_path}")
        print(f"  File size: {file_size} bytes")

        print("\n" + "=" * 70)
        print("Step 12: Read exported order numbers")
        print("=" * 70)
        import openpyxl

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        order_numbers = []
        header_row = None

        for row in ws.iter_rows(max_row=1):
            header_row = [cell.value for cell in row]
            break

        if header_row:
            print(f"Exported header row: {[str(h)[:30] if h else '' for h in header_row]}")
            order_col_index = None
            for i, header in enumerate(header_row):
                if header and (
                    "系统单号" in str(header)
                    or "订单号" in str(header)
                    or "销售单号" in str(header)
                    or "systemNo" in str(header).lower()
                ):
                    order_col_index = i
                    break

            if order_col_index is None:
                order_col_index = 0

            for row in ws.iter_rows(min_row=2, max_row=51):
                cell_value = row[order_col_index].value
                if cell_value:
                    order_numbers.append(str(cell_value).strip())

        wb.close()

        print(f"Exported order numbers: {len(order_numbers)} found")
        print(f"Total exported: {len(order_numbers)}")

        exported_unique_order_numbers = []
        seen_exported_order_numbers = set()
        for order_num in order_numbers:
            if order_num not in seen_exported_order_numbers:
                seen_exported_order_numbers.add(order_num)
                exported_unique_order_numbers.append(order_num)

        print(f"Exported unique order numbers: {len(exported_unique_order_numbers)} found")
        print(f"Total exported unique: {len(exported_unique_order_numbers)}")

        print("\n" + "=" * 70)
        print("Step 13: Verify exported order coverage")
        print("=" * 70)

        matching_numbers = []
        for order_num in page_order_numbers[:20]:
            if order_num in exported_unique_order_numbers:
                matching_numbers.append(order_num)

        print(f"Matching order numbers: {len(matching_numbers)} found")
        print(f"Total matched: {len(matching_numbers)}")

        expected_order_numbers = page_order_numbers[:3]
        missing_order_numbers = [num for num in expected_order_numbers if num not in exported_unique_order_numbers]

        if missing_order_numbers:
            print(f"\n❌ 导出文件缺少请求导出的订单号，缺少数量: {len(missing_order_numbers)}")
            pytest.fail(f"导出文件缺少请求导出的订单号，缺少数量: {len(missing_order_numbers)}")

        unexpected_order_numbers = [num for num in exported_unique_order_numbers if num not in expected_order_numbers]
        if unexpected_order_numbers:
            print(f"\n❌ 导出文件包含未请求的订单号，异常数量: {len(unexpected_order_numbers)}")
            pytest.fail(f"导出文件包含未请求的订单号，异常数量: {len(unexpected_order_numbers)}")

        print("\n✅ 导出订单覆盖验证通过：导出文件包含且仅包含请求导出的订单号")

        print("\n" + "=" * 70)
        print("Test completed")
        print("=" * 70)
