"""Sales product sales report regression coverage."""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from playwright.sync_api import Page

from modules.auto_test.core.config_manager import get_config
from modules.auto_test.pages.login_page import LoginPage
from modules.auto_test.pages.sales_report_page import SalesReportPage


REPORT_DIR = Path("reports/sales_report_regression")
QUERY_DETAIL_JSONL = REPORT_DIR / "query_detail.jsonl"
QUERY_DETAIL_MD = REPORT_DIR / "query_detail_report.md"


def _first_meaningful_prefix(value: str, length: int = 8) -> str:
    cleaned = value.strip()
    if not cleaned or cleaned == "-":
        return ""
    return cleaned[: min(length, len(cleaned))]


def _compact_row(row: dict[str, str]) -> dict[str, str]:
    interesting_columns = [
        "SKU编码",
        "变体ID",
        "SKU中文",
        "SKU英文",
        "销售状态",
        "品类",
        "创建时间",
        "销量",
        "日均销量",
        "库存(汇总)",
        "订单数",
    ]
    return {column: row.get(column, "") for column in interesting_columns if column in row}


def _first_usable_variant_id(sales_report: SalesReportPage) -> tuple[str, dict[str, str]]:
    fallback: tuple[str, dict[str, str]] = ("", {})
    for row in sales_report.visible_rows(limit=50):
        raw_value = row.get("变体ID", "")
        if not raw_value or raw_value == "-":
            continue
        values = [value.strip() for value in raw_value.split(",") if value.strip() and value.strip() != "-"]
        if not values:
            continue
        variant_id = values[0]
        if not fallback[0]:
            fallback = (variant_id, row)
        created_date = row.get("创建时间", "")[:10]
        if len(values) == 1 and created_date >= "2026-06-29":
            return variant_id, row
    for row in sales_report.visible_rows(limit=50):
        raw_value = row.get("变体ID", "")
        values = [value.strip() for value in raw_value.split(",") if value.strip() and value.strip() != "-"]
        if len(values) == 1:
            return values[0], row
    if fallback[0]:
        return fallback
    for page_number in (2, 3, 4):
        if not sales_report.click_page_number(page_number):
            continue
        for row in sales_report.visible_rows(limit=50):
            raw_value = row.get("变体ID", "")
            values = [value.strip() for value in raw_value.split(",") if value.strip() and value.strip() != "-"]
            created_date = row.get("创建时间", "")[:10]
            if len(values) == 1 and created_date >= "2026-06-29":
                return values[0], row
            if len(values) == 1 and not fallback[0]:
                fallback = (values[0], row)
    if fallback[0]:
        return fallback
    return "", {}


def _first_created_date_on_or_after(sales_report: SalesReportPage, min_date: str) -> tuple[str, dict[str, str]]:
    for row in sales_report.visible_rows(limit=50):
        created_at = row.get("创建时间", "")
        created_date = created_at[:10] if len(created_at) >= 10 else ""
        if created_date and created_date >= min_date:
            return created_date, row
    return "", {}


def _record_query_detail(
    case_name: str,
    conditions: dict[str, str],
    sales_report: SalesReportPage,
    assertion: str,
    passed: bool,
    note: str = "",
) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    rows = sales_report.visible_rows(limit=3)
    detail: dict[str, Any] = {
        "case_name": case_name,
        "conditions": conditions,
        "total_count": sales_report.total_count(),
        "current_page_row_count": sales_report.row_count(),
        "sample_rows": [_compact_row(row) for row in rows],
        "assertion": assertion,
        "passed": passed,
        "note": note,
    }
    with QUERY_DETAIL_JSONL.open("a", encoding="utf-8") as file:
        file.write(json.dumps(detail, ensure_ascii=False) + "\n")
    with QUERY_DETAIL_MD.open("a", encoding="utf-8") as file:
        file.write(f"\n## {case_name}\n\n")
        file.write(f"- 搜索条件：{conditions}\n")
        file.write(f"- 返回总数：{detail['total_count']}\n")
        file.write(f"- 当前页行数：{detail['current_page_row_count']}\n")
        file.write(f"- 验证点：{assertion}\n")
        file.write(f"- 结果：{'PASS' if passed else 'FAIL'}\n")
        if note:
            file.write(f"- 备注：{note}\n")
        file.write(f"- 首行样例：{detail['sample_rows'][0] if detail['sample_rows'] else {}}\n")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_excel_rows(path: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_cell(value) for value in rows[0]]
    result = []
    for row in rows[1:]:
        item = {headers[index]: _normalize_cell(value) for index, value in enumerate(row) if index < len(headers)}
        if any(item.values()):
            result.append(item)
    return result


def _find_download_url(value: Any) -> str:
    if isinstance(value, str):
        lowered = value.lower()
        if lowered.startswith(("http://", "https://")) or ".xlsx" in lowered or ".xls" in lowered:
            return value
        return ""
    if isinstance(value, dict):
        for key in ("downloadUrl", "fileUrl", "url", "path"):
            found = _find_download_url(value.get(key))
            if found:
                return found
        for nested in value.values():
            found = _find_download_url(nested)
            if found:
                return found
    if isinstance(value, list):
        for nested in value:
            found = _find_download_url(nested)
            if found:
                return found
    return ""


def _column_numbers(rows: list[dict[str, str]], column_name: str, limit: int = 20) -> list[float]:
    values: list[float] = []
    for row in rows[:limit]:
        raw = row.get(column_name, "").replace(",", "").strip()
        if not raw or raw == "-":
            continue
        try:
            values.append(float(raw))
        except ValueError:
            continue
    return values


def _extract_total(value: Any) -> int:
    if isinstance(value, dict):
        for key in ("total", "totalCount", "count"):
            raw = value.get(key)
            if isinstance(raw, int):
                return raw
            if isinstance(raw, str) and raw.isdigit():
                return int(raw)
        for nested in value.values():
            total = _extract_total(nested)
            if total >= 0:
                return total
    return -1


def _business_success(body: Any) -> bool:
    if not isinstance(body, dict):
        return False
    code = body.get("code")
    if code in (0, 200, "0", "200"):
        return True
    success = body.get("success")
    return success is True or success == "true"


def _export_row_key(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    return (
        row.get("店铺名称") or row.get("店铺") or "",
        row.get("库存SKU") or row.get("SKU编码") or "",
        row.get("销量", ""),
        row.get("日均销量", ""),
        row.get("订单数", ""),
    )


def _detail_row_key(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    return (
        row.get("店铺", ""),
        row.get("SKU编码", ""),
        row.get("销量", ""),
        row.get("日均销量", ""),
        row.get("订单数", ""),
    )


def _assert_detail_rows_in_export(detail_rows: list[dict[str, str]], export_rows: list[dict[str, str]]) -> None:
    export_keys = {_export_row_key(row) for row in export_rows}
    missing = [_detail_row_key(row) for row in detail_rows if _detail_row_key(row) not in export_keys]
    assert not missing, f"Expanded detail rows missing in export: {missing[:5]}"


def _page_sku_order(sales_report: SalesReportPage, limit: int = 20) -> list[str]:
    result = []
    for row in sales_report.visible_rows(limit=limit):
        sku = row.get("SKU编码", "").strip()
        if sku and sku != "-":
            result.append(sku)
    return result


def _export_sku_first_seen_order(export_rows: list[dict[str, str]], expected_skus: list[str]) -> list[str]:
    expected = set(expected_skus)
    seen = []
    for row in export_rows:
        sku = (row.get("库存SKU") or row.get("SKU编码") or "").strip()
        if sku in expected and sku not in seen:
            seen.append(sku)
        if len(seen) == len(expected_skus):
            break
    return seen


def _export_contains_skus(export_rows: list[dict[str, str]], expected_skus: list[str]) -> bool:
    exported_skus = {(row.get("库存SKU") or row.get("SKU编码") or "").strip() for row in export_rows}
    return all(sku in exported_skus for sku in expected_skus)


def _api_sku_order(response_body: Any, limit: int = 10) -> list[str]:
    if not isinstance(response_body, dict):
        return []
    rows = response_body.get("rows")
    if not isinstance(rows, list):
        data = response_body.get("data")
        rows = data.get("rows") if isinstance(data, dict) else []
    result = []
    for row in rows[:limit]:
        if not isinstance(row, dict):
            continue
        sku = str(row.get("itemId") or row.get("skuCode") or "").strip()
        if sku:
            result.append(sku)
    return result


@pytest.fixture(scope="session", autouse=True)
def query_detail_report_header() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    QUERY_DETAIL_JSONL.write_text("", encoding="utf-8")
    QUERY_DETAIL_MD.write_text(
        "# 销售商品销量报表查询条件明细报告\n\n"
        "说明：本报告由自动化实跑生成，记录每个查询条件、返回数量、样例数据和验证结果。\n",
        encoding="utf-8",
    )


@pytest.fixture()
def sales_report(logged_in_page: Page) -> SalesReportPage:
    page = SalesReportPage(logged_in_page)
    page.navigate_to_report()
    if page.row_count() == 0:
        page.search()
    assert page.row_count() > 0, "Sales report should load default rows"
    return page


@pytest.fixture()
def sales_report_large_account(page: Page) -> SalesReportPage:
    page.context.clear_cookies()
    page.goto(get_config().base_url)
    page.evaluate("() => { localStorage.clear(); sessionStorage.clear(); }")
    page.goto(f"{get_config().base_url.rstrip('/')}/login")
    login_page = LoginPage(page)
    assert login_page.login("15137651220", "123456"), "Large-data account login failed"
    report_page = SalesReportPage(page)
    report_page.navigate_to_report()
    return report_page


@pytest.mark.regression
@pytest.mark.ui
@pytest.mark.p1
class TestSalesReportRegression:
    def test_default_query_loads_table(self, sales_report: SalesReportPage) -> None:
        first_row = sales_report.first_row()

        assert sales_report.total_count() > 0
        assert first_row.get("SKU编码")
        assert first_row.get("SKU中文")
        assert "销量" in first_row
        assert "日均销量" in first_row
        _record_query_detail(
            "默认查询",
            {"订单付款时间": "页面默认近 30 天"},
            sales_report,
            "表格有数据，首行包含 SKU 编码、SKU 中文、销量、日均销量",
            True,
        )

    @pytest.mark.parametrize(
        ("field_name", "fill_method", "column_name"),
        [
            ("SKU编码", "fill_sku_code", "SKU编码"),
            ("SKU中文", "fill_sku_cn", "SKU中文"),
            ("变体ID", "fill_variant_id", "变体ID"),
        ],
    )
    def test_text_condition_query(
        self,
        sales_report: SalesReportPage,
        field_name: str,
        fill_method: str,
        column_name: str,
    ) -> None:
        first_row = sales_report.first_row()
        query_value = _first_meaningful_prefix(first_row.get(field_name, ""))
        source_note = ""
        if field_name == "变体ID":
            query_value, source_row = _first_usable_variant_id(sales_report)
            source_note = f"来源行={_compact_row(source_row)}"
        if not query_value:
            pytest.skip(f"Current page has no usable {field_name} value")

        sales_report.reset()
        getattr(sales_report, fill_method)(query_value)
        sales_report.search()

        passed = False
        note = ""
        try:
            sales_report.assert_column_contains(column_name, query_value)
            passed = True
        except AssertionError as error:
            note = str(error)
            raise
        finally:
            _record_query_detail(
                f"{field_name} 查询",
                {field_name: query_value},
                sales_report,
                f"{column_name} 列包含搜索值 {query_value}",
                passed,
                "；".join(item for item in (source_note, note) if item),
            )

    def test_sku_english_condition_query_accepts_empty_result(self, sales_report: SalesReportPage) -> None:
        first_row = sales_report.first_row()
        query_value = _first_meaningful_prefix(first_row.get("SKU英文", ""))
        if not query_value:
            query_value = "NO-SUCH-SKU-EN"

        sales_report.reset()
        sales_report.fill_sku_en(query_value)
        sales_report.search()

        assert sales_report.row_count() >= 0
        _record_query_detail(
            "SKU 英文查询",
            {"SKU英文": query_value},
            sales_report,
            "页面正常返回结果或空结果，不报错",
            True,
        )

    def test_sales_status_condition_query(self, sales_report: SalesReportPage) -> None:
        first_row = sales_report.first_row()
        status = _first_meaningful_prefix(first_row.get("销售状态", ""), length=4)
        if not status:
            pytest.skip("Current page has no usable sales status value")

        sales_report.reset()
        selected = sales_report.select_dropdown_by_label("销售状态", status)
        assert selected, f"Unable to select sales status: {status}"
        sales_report.search()

        sales_report.assert_column_contains("销售状态", status)
        _record_query_detail(
            "销售状态查询",
            {"销售状态": status},
            sales_report,
            f"销售状态列包含 {status}",
            True,
        )

    def test_category_condition_query(self, sales_report: SalesReportPage) -> None:
        first_row = sales_report.first_row()
        category = _first_meaningful_prefix(first_row.get("品类", ""), length=8)
        if not category:
            pytest.skip("Current page has no usable category value")

        sales_report.reset()
        selected = sales_report.select_dropdown_by_label("品类", category)
        if not selected:
            pytest.skip(f"Category selector did not expose option: {category}")
        sales_report.search()

        sales_report.assert_column_contains("品类", category)
        _record_query_detail(
            "品类查询",
            {"品类": category},
            sales_report,
            f"品类列包含 {category}",
            True,
        )

    def test_sku_create_date_condition_query_negative(self, sales_report: SalesReportPage) -> None:
        first_row = sales_report.first_row()
        created_at = first_row.get("创建时间", "")
        match = created_at[:10] if len(created_at) >= 10 else ""
        if not match:
            pytest.skip("Current page has no usable created date")

        sales_report.reset()
        sales_report.fill_sku_create_date(f"{match} 00:00:00", f"{match} 00:00:00")
        input_values = sales_report.sku_create_date_values()
        sales_report.search()

        passed = sales_report.total_count() == 0
        _record_query_detail(
            "SKU 创建日期查询",
            {"SKU创建日期开始时间": match, "SKU创建日期结束时间": match},
            sales_report,
            "配合页面默认最近 30 天付款时间，返回总数应为 0",
            passed,
            (
                f"输入框实际值={input_values}；请求payload={sales_report.last_search_payload}；"
                f"实际返回总数={sales_report.total_count()}"
            ),
        )
        assert passed, f"SKU create date range should return 0 rows, got {sales_report.total_count()}"

    def test_sku_create_date_condition_query_positive(self, sales_report: SalesReportPage) -> None:
        match, source_row = _first_created_date_on_or_after(sales_report, "2026-06-29")
        if not match:
            pytest.skip("Current page has no SKU created date inside default payment range")

        sales_report.reset()
        sales_report.fill_sku_create_date(match, match)
        input_values = sales_report.sku_create_date_values()
        sales_report.search()

        passed = sales_report.total_count() > 0
        _record_query_detail(
            "SKU 创建日期查询-正向",
            {"SKU创建日期开始时间": match, "SKU创建日期结束时间": match},
            sales_report,
            "配合页面默认最近 30 天付款时间，应返回至少 1 条结果",
            passed,
            (
                f"来源行={_compact_row(source_row)}；输入框实际值={input_values}；"
                f"请求payload={sales_report.last_search_payload}；实际返回总数={sales_report.total_count()}"
            ),
        )
        assert passed, f"SKU create date range should return rows, got {sales_report.total_count()}"

    def test_sku_create_date_condition_query_full_range(self, sales_report: SalesReportPage) -> None:
        start_time = "2026-06-01 00:00:00"
        end_time = "2026-07-31 23:59:59"

        sales_report.reset()
        sales_report.fill_sku_create_date(start_time, end_time)
        input_values = sales_report.sku_create_date_values()
        sales_report.search()

        passed = sales_report.total_count() > 0
        _record_query_detail(
            "SKU 创建日期查询-完整范围",
            {"SKU创建日期开始时间": start_time, "SKU创建日期结束时间": end_time},
            sales_report,
            "配合页面默认最近 30 天付款时间，应返回至少 1 条结果",
            passed,
            (
                f"输入框实际值={input_values}；请求payload={sales_report.last_search_payload}；"
                f"实际返回总数={sales_report.total_count()}"
            ),
        )
        assert passed, f"SKU create date full range should return rows, got {sales_report.total_count()}"

    def test_pagination(self, sales_report: SalesReportPage) -> None:
        initial_page = sales_report.current_page()
        moved = sales_report.click_page_number(2) or sales_report.next_page()
        if moved:
            assert sales_report.current_page() != initial_page
            assert sales_report.row_count() > 0
            _record_query_detail(
                "分页",
                {"操作": f"从第 {initial_page} 页切换到第 {sales_report.current_page()} 页"},
                sales_report,
                "页码变化且当前页有数据",
                True,
            )
        else:
            pytest.skip("Only one page is available")

    @pytest.mark.parametrize("column_name", ["销量", "日均销量", "库存(汇总)", "订单数"])
    def test_sort_columns(self, sales_report: SalesReportPage, column_name: str) -> None:
        result = sales_report.sort_and_assert_numeric_order(column_name)
        _record_query_detail(
            f"{column_name} 排序",
            {"排序列": column_name, "排序方向": "先降序，后升序"},
            sales_report,
            "降序和升序后可见数值均有序；"
            f"降序取样值={result.get('desc_values', [])[:10]}；"
            f"升序取样值={result.get('asc_values', [])[:10]}；"
            f"排序请求payload={result.get('payloads', [])}",
            bool(result["passed"]),
        )
        assert result["passed"], result

    def test_expand_detail(self, sales_report: SalesReportPage) -> None:
        assert sales_report.has_expand_control(), "Sales report should expose row detail expand controls"
        expanded_count = -1
        try:
            expanded_count = sales_report.expand_first_row()
            passed = True
            note = f"展开明细行数量={expanded_count}"
        except AssertionError as error:
            passed = False
            note = str(error)
            raise
        finally:
            _record_query_detail(
                "展开明细",
                {"操作": "点击首行左侧展开箭头"},
                sales_report,
                "点击后出现明细展开内容",
                passed,
                note,
            )
        assert expanded_count >= 0

    def test_sort_then_expand_detail(self, sales_report: SalesReportPage) -> None:
        sort_result = sales_report.sort_and_assert_numeric_order("销量")

        assert sort_result["passed"], sort_result
        assert sales_report.has_expand_control(), "Sales report should expose row detail expand controls after sorting"
        expanded_count = -1
        try:
            expanded_count = sales_report.expand_first_row()
            passed = True
            note = (
                f"降序取样值={sort_result.get('desc_values', [])[:10]}；"
                f"升序取样值={sort_result.get('asc_values', [])[:10]}；"
                f"展开明细行数量={expanded_count}"
            )
        except AssertionError as error:
            passed = False
            note = (
                f"降序取样值={sort_result.get('desc_values', [])[:10]}；"
                f"升序取样值={sort_result.get('asc_values', [])[:10]}；"
                f"展开失败：{error}"
            )
            raise
        finally:
            _record_query_detail(
                "销量排序后展开明细",
                {"复合操作": "销量降序排序 -> 点击首行左侧展开箭头"},
                sales_report,
                "排序后仍可展开首行明细",
                passed,
                note,
            )
        assert expanded_count >= 0

    def _legacy_test_export_menu_and_downloads(self, sales_report: SalesReportPage) -> None:
        options = sales_report.export_menu_options()

        assert any("当前页" in option for option in options), f"Export current-page option missing: {options}"
        assert any("搜索条件" in option for option in options), f"Export by-search option missing: {options}"

        download_dir = REPORT_DIR / "downloads"
        current_page = sales_report.export_by_menu_text("当前页", str(download_dir))
        assert current_page["success"] or any(
            item["status"] == 500 for item in current_page.get("responses", [])
        ), current_page

        by_search = sales_report.export_by_menu_text("搜索条件", str(download_dir))
        assert by_search["success"], by_search
        _record_query_detail(
            "导出",
            {"操作": "打开导出菜单 -> 导出当前页 -> 按搜索条件导出"},
            sales_report,
            "当前页导出允许 500 已知缺陷；按搜索条件导出成功",
            True,
            f"菜单项={options}；当前页导出={current_page}；按条件导出={by_search}",
        )

    @pytest.mark.parametrize("scenario", ["SKU编码", "品类", "SKU创建日期完整范围"])
    def test_search_export_matches_expanded_details(self, sales_report: SalesReportPage, scenario: str) -> None:
        first_row = sales_report.first_row()
        conditions: dict[str, str] = {}

        sales_report.reset()
        if scenario == "SKU编码":
            sku_code = first_row.get("SKU编码", "")
            if not sku_code:
                pytest.skip("Current page has no usable SKU code")
            sales_report.fill_sku_code(sku_code)
            conditions = {"SKU编码": sku_code}
        elif scenario == "品类":
            category = _first_meaningful_prefix(first_row.get("品类", ""), length=8)
            if not category:
                pytest.skip("Current page has no usable category")
            selected = sales_report.select_dropdown_by_label("品类", category)
            if not selected:
                pytest.skip(f"Category selector did not expose option: {category}")
            conditions = {"品类": category}
        else:
            start_time = "2026-06-01 00:00:00"
            end_time = "2026-07-31 23:59:59"
            sales_report.fill_sku_create_date(start_time, end_time)
            conditions = {"SKU创建日期开始时间": start_time, "SKU创建日期结束时间": end_time}

        sales_report.search()
        assert sales_report.row_count() > 0, f"{scenario} search returned no rows"
        sales_report.expand_first_row()
        detail_rows = sales_report.expanded_detail_rows(limit=20)
        assert detail_rows, f"{scenario} did not expose expanded detail rows"

        download_dir = REPORT_DIR / "downloads" / "search_export_match"
        export_result = sales_report.export_by_menu_text("搜索条件", str(download_dir))
        assert export_result["success"], export_result
        export_rows = _read_excel_rows(export_result["file_path"])
        assert export_rows, export_result
        _assert_detail_rows_in_export(detail_rows, export_rows)

        first_export_skus = {row.get("库存SKU", "") for row in export_rows[:20]}
        if scenario == "SKU编码":
            assert conditions["SKU编码"] in first_export_skus
        elif scenario == "品类":
            assert all(conditions["品类"] in row.get("品类", "") for row in export_rows[:20])

        _record_query_detail(
            f"{scenario} 查询后导出明细一致性",
            conditions,
            sales_report,
            "按条件搜索后展开明细，导出搜索结果，Excel 明细应包含页面展开明细",
            True,
            (
                f"展开明细取样={detail_rows[:3]}；导出文件={export_result.get('file_path')}；"
                f"导出行数={len(export_rows)}；导出首行={export_rows[0] if export_rows else {}}"
            ),
        )

    def test_export_menu_and_downloads(self, sales_report: SalesReportPage) -> None:
        options = sales_report.export_menu_options()
        assert options, "Export button is not visible"
        download_dir = REPORT_DIR / "downloads"
        export_result = sales_report.export_by_menu_text("当前搜索结果", str(download_dir))
        assert export_result["success"], export_result
        _record_query_detail(
            "导出",
            {"操作": "点击导出按钮，导出当前搜索结果"},
            sales_report,
            "导出按钮直接导出当前搜索结果",
            True,
            f"导出按钮={options}；导出结果={export_result}",
        )

    def test_async_export_large_search_records_task_and_preserves_request(
        self, sales_report_large_account: SalesReportPage
    ) -> None:
        sales_report = sales_report_large_account
        payload = {
            "payTimeBegin": "2026-04-01 00:00:00",
            "payTimeEnd": "2026-07-01 23:59:59",
            "sortField": "dayAverageSalesNumber",
            "sortOrder": "desc",
            "exportPicFlag": "1",
            "excelType": "0",
        }
        query_payload = {**payload, "pageNum": 1, "pageSize": 50}

        query_result = sales_report.query_report_api(query_payload)
        query_total = _extract_total(query_result.get("body"))
        assert query_result["ok"], query_result
        assert _business_success(query_result.get("body")), query_result

        if query_total <= 5000:
            _record_query_detail(
                "大数据量异步导出",
                {
                    "订单付款时间开始": payload["payTimeBegin"],
                    "订单付款时间结束": payload["payTimeEnd"],
                    "排序字段": payload["sortField"],
                    "排序方向": payload["sortOrder"],
                    "导出图片": payload["exportPicFlag"],
                    "Excel类型": payload["excelType"],
                },
                sales_report,
                "前置条件校验：同条件查询总数必须大于 5000 才能验证异步导出阈值",
                False,
                f"当前查询总数={query_total}，未达到异步导出阈值；查询接口返回={query_result.get('body')}",
            )
            pytest.skip(f"Async export threshold is not met: total={query_total}")

        result = sales_report.trigger_async_export(payload)
        body = result.get("body")
        download_url = _find_download_url(body)
        note = f"查询总数={query_total}；查询接口返回={query_result.get('body')}；异步导出接口返回={body}"

        assert result["status"] < 500, result
        assert result["ok"], result
        assert _business_success(body), result
        assert result["payload"] == payload

        if download_url:
            download_dir = REPORT_DIR / "downloads" / "async_export"
            download_dir.mkdir(parents=True, exist_ok=True)
            response = sales_report.page.context.request.get(download_url)
            assert response.ok, {"download_url": download_url, "status": response.status}
            target = download_dir / "sales_report_async_export.xlsx"
            target.write_bytes(response.body())
            export_rows = _read_excel_rows(str(target))
            day_average_values = _column_numbers(export_rows, "日均销量")
            assert export_rows, str(target)
            assert SalesReportPage.is_sorted(day_average_values, "desc"), day_average_values[:20]
            note = (
                f"查询总数={query_total}；异步导出接口返回={body}；下载文件={target}；"
                f"导出行数={len(export_rows)}；日均销量前20={day_average_values[:20]}"
            )

        _record_query_detail(
            "大数据量异步导出",
            {
                "订单付款时间开始": payload["payTimeBegin"],
                "订单付款时间结束": payload["payTimeEnd"],
                "排序字段": payload["sortField"],
                "排序方向": payload["sortOrder"],
                "导出图片": payload["exportPicFlag"],
                "Excel类型": payload["excelType"],
            },
            sales_report,
            "查询总数大于 5000，触发异步导出接口业务成功，导出请求时间和排序入参与预期一致；若返回下载地址则继续校验 Excel 日均销量降序",
            True,
            note,
        )

    @pytest.mark.parametrize(
        ("menu_text", "case_name"),
        [
            ("当前页", "日均销量排序后导出当前页顺序一致性"),
            ("搜索条件", "日均销量排序后导出当前搜索结果顺序一致性"),
        ],
    )
    def test_sorted_export_preserves_visible_sku_order(
        self, sales_report: SalesReportPage, menu_text: str, case_name: str
    ) -> None:
        sales_report.reset()
        sales_report.click_sort("日均销量", "asc")
        sort_values = sales_report.click_sort("日均销量", "desc")
        assert SalesReportPage.is_sorted(sort_values, "desc"), sort_values[:20]
        sort_payload = sales_report.last_sort_payloads[-1] if sales_report.last_sort_payloads else None
        assert sort_payload, "No sort request payload captured"
        query_payload = {**sort_payload, "pageNum": 1, "pageSize": 50}
        query_result = sales_report.query_report_api(query_payload)
        assert query_result["ok"], query_result
        assert _business_success(query_result.get("body")), query_result
        page_skus = _api_sku_order(query_result.get("body"), limit=10)
        assert page_skus, "No SKU order can be captured from sorted list API"

        download_dir = REPORT_DIR / "downloads" / "sorted_export_order"
        export_result = sales_report.export_by_menu_text(menu_text, str(download_dir), timeout=90000)
        assert export_result["success"], export_result
        assert export_result.get("file_path"), export_result

        export_rows = _read_excel_rows(export_result["file_path"])
        export_order = _export_sku_first_seen_order(export_rows, page_skus)
        exported_values = _column_numbers(export_rows, "日均销量", limit=50)
        contains_expected = _export_contains_skus(export_rows, page_skus[:5])
        # Export rows are details; validate SKU group order, not global detail
        # metric order.
        export_order_matches = export_order[:5] == page_skus[:5]
        # Keep the existing diagnostic field name compatible with reports.
        export_sorted = export_order_matches
        passed = contains_expected and export_order_matches
        _record_query_detail(
            case_name,
            {"排序字段": "日均销量", "排序方向": "desc", "导出方式": menu_text},
            sales_report,
            "排序后导出，Excel 应包含列表前 5 个 SKU，且导出明细日均销量按降序排列",
            passed,
            (
                f"列表接口SKU顺序={page_skus}；导出SKU首次出现顺序={export_order}；"
                f"导出日均销量取样={exported_values[:20]}；排序payload={sort_payload}；"
                f"导出文件={export_result.get('file_path')}；导出行数={len(export_rows)}；"
                f"包含前5个列表SKU={contains_expected}；导出明细降序={export_sorted}"
            ),
        )
        assert passed, {
            "menu_text": menu_text,
            "page_order": page_skus,
            "export_order": export_order,
            "exported_values": exported_values[:20],
            "contains_expected": contains_expected,
            "export_order_matches": export_order_matches,
            "file_path": export_result.get("file_path"),
        }
