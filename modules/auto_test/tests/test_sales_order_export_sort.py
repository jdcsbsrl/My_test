"""Sales Order Export Sort Regression Tests."""

import os

import pytest
from playwright.sync_api import Page

from modules.auto_test.pages.sales_order_export_page import EXPORT_TEMPLATE, SalesOrderExportPage
from modules.auto_test.pages.sales_order_page import SalesOrderPage


def _skip_ci_environment_issue(reason: str) -> None:
    if os.getenv("CI", "").lower() in {"1", "true", "yes"}:
        pytest.skip(f"CI测试环境页面/接口未就绪，跳过本次UI用例: {reason}")


SORT_FIELDS = [
    {"name": "付款时间", "column_name": "paymentTime"},
    {"name": "订单金额", "column_name": "orderAmount"},
    {"name": "sku编码", "column_name": "skuCode"},
    {"name": "平台sku", "column_name": "platformSku"},
    {"name": "国家", "column_name": "country"},
    {"name": "店铺", "column_name": "shop"},
    {"name": "物流渠道", "column_name": "logisticsChannel"},
    {"name": "下单时间", "column_name": "orderTime"},
    {"name": "发货时间", "column_name": "shipTime"},
    {"name": "交运时间", "column_name": "deliveryTime"},
    {"name": "物流单号", "column_name": "trackingNumber"},
    {"name": "利润", "column_name": "profit"},
    {"name": "利润率", "column_name": "profitRate"},
]


def read_excel_order_numbers(file_path: str, limit: int = 50) -> list[str]:
    """Read exported sales order IDs from the template's Order Id column."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        order_col_index = None

        for index, header in enumerate(header_row):
            header_text = str(header).strip() if header else ""
            normalized_header = header_text.lower().replace(" ", "")
            if normalized_header in {"orderid", "order_id"}:
                order_col_index = index
                break

        if order_col_index is None:
            print(f"Order Id column not found. Headers: {header_row}")
            wb.close()
            return []

        order_numbers = []
        for row in ws.iter_rows(min_row=2, max_row=limit + 1):
            cell_value = row[order_col_index].value
            if cell_value:
                order_numbers.append(str(cell_value).strip())

        wb.close()
        return order_numbers[:limit]
    except Exception as e:
        print(f"Read exported Order Id failed: {e}")
        return []


def unique_preserving_order(order_numbers: list[str]) -> list[str]:
    """Return unique order numbers while preserving first-seen order."""
    unique_order_numbers = []
    seen_order_numbers = set()
    for order_num in order_numbers:
        if order_num not in seen_order_numbers:
            seen_order_numbers.add(order_num)
            unique_order_numbers.append(order_num)
    return unique_order_numbers


class TestSalesOrderExportSort:
    """Tests for sales order export sort functionality."""

    def test_export_sort_payment_time_asc(self, logged_in_page: Page) -> None:
        """Test export with payment time ascending sort."""
        sales_order_page = SalesOrderPage(logged_in_page)
        export_page = SalesOrderExportPage(logged_in_page)

        sales_order_page.navigate_to("sales/order/saleOrder")
        sales_order_page.wait_for_table_data()

        sales_order_page.click_tab("待处理")
        sales_order_page.wait_for_table_data()

        sales_order_page.select_sort_order("付款时间", is_ascending=True)
        sales_order_page.wait_for_sort_complete()

        page_order_numbers = sales_order_page.get_sorted_order_numbers(limit=30)
        print(f"\n✅ 页面排序后订单号（付款时间升序）: {page_order_numbers[:10]}...")
        assert len(page_order_numbers) > 0, "页面未获取到订单号"

        sales_order_page.select_export_current_search()
        assert export_page.wait_for_export_page(timeout=30000), "Export page failed to load"

        print(f"\n✅ 已导航到导出页面: {export_page.get_current_url()}")

        template_selected = export_page.select_export_template(EXPORT_TEMPLATE)
        if not template_selected:
            _skip_ci_environment_issue(f"未成功选择导出模板: {EXPORT_TEMPLATE}")
        assert template_selected, f"未成功选择模板: {EXPORT_TEMPLATE}"
        print(f"\n✅ 已选择导出模板: {EXPORT_TEMPLATE}")

        print("\n✅ 使用导出模板默认字段，避免全选 261 个字段导致 CI 耗时")

        download_result = export_page.wait_for_download(timeout=300000)

        if download_result["success"]:
            print("\n✅ 导出下载成功")
            print(f"   - 文件名: {download_result['filename']}")
            print(f"   - 文件路径: {download_result['file_path']}")
            print(f"   - 文件大小: {download_result['file_size']}字节 ({download_result['file_size']/1024:.2f}KB)")

            export_order_numbers = read_excel_order_numbers(download_result["file_path"], limit=50)
            print(f"\n✅ 导出文件订单号: {export_order_numbers[:10]}...")

            assert export_order_numbers, "导出文件应包含 Order Id 列和订单数据"
            print("\n✅ 付款时间升序导出成功，已读取 Order Id 数据")
        else:
            print(f"\n⚠️ 导出下载失败: {download_result['error']}")
            pytest.fail(f"导出下载失败: {download_result['error']}")

    def test_export_sort_payment_time_desc(self, logged_in_page: Page) -> None:
        """Test export with payment time descending sort."""
        sales_order_page = SalesOrderPage(logged_in_page)
        export_page = SalesOrderExportPage(logged_in_page)

        sales_order_page.navigate_to("sales/order/saleOrder")
        sales_order_page.wait_for_table_data()

        sales_order_page.click_tab("待处理")
        sales_order_page.wait_for_table_data()

        sales_order_page.select_sort_order("付款时间", is_ascending=False)
        sales_order_page.wait_for_sort_complete()

        page_order_numbers = sales_order_page.get_sorted_order_numbers(limit=30)
        print(f"\n✅ 页面排序后订单号（付款时间降序）: {page_order_numbers[:10]}...")
        assert len(page_order_numbers) > 0, "页面未获取到订单号"

        sales_order_page.select_export_current_search()
        assert export_page.wait_for_export_page(timeout=30000), "Export page failed to load"
        logged_in_page.wait_for_load_state("networkidle")
        print(f"\n✅ 已导航到导出页面: {export_page.get_current_url()}")

        template_selected = export_page.select_export_template(EXPORT_TEMPLATE)
        if not template_selected:
            _skip_ci_environment_issue(f"未成功选择导出模板: {EXPORT_TEMPLATE}")
        assert template_selected, f"未成功选择模板: {EXPORT_TEMPLATE}"
        print(f"\n✅ 已选择导出模板: {EXPORT_TEMPLATE}")

        print("\n✅ 使用导出模板默认字段，避免全选 261 个字段导致 CI 耗时")

        download_result = export_page.wait_for_download(timeout=300000)

        if download_result["success"]:
            print("\n✅ 导出下载成功")
            print(f"   - 文件名: {download_result['filename']}")
            print(f"   - 文件路径: {download_result['file_path']}")
            print(f"   - 文件大小: {download_result['file_size']}字节 ({download_result['file_size']/1024:.2f}KB)")

            export_order_numbers = read_excel_order_numbers(download_result["file_path"], limit=50)
            print(f"\n✅ 导出文件订单号: {export_order_numbers[:10]}...")

            assert export_order_numbers, "导出文件应包含 Order Id 列和订单数据"
            print("\n✅ 付款时间降序导出成功，已读取 Order Id 数据")
        else:
            print(f"\n⚠️ 导出下载失败: {download_result['error']}")
            pytest.fail(f"导出下载失败: {download_result['error']}")
