"""Quick test to verify export flow with system order number."""

import os
import sys
import time

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

from playwright.sync_api import Page
import pytest

from modules.auto_test.core.config_manager import get_config
from modules.auto_test.facades.sales_order_facade import SalesOrderFacade
from modules.auto_test.pages.sales_order_export_page import EXPORT_TEMPLATE, SalesOrderExportPage


@pytest.mark.regression
@pytest.mark.ui
def test_export_with_system_order_no(logged_in_page: Page) -> None:
    """Test export with system order number verification."""
    page = logged_in_page
    facade = SalesOrderFacade(page)

    print("\n" + "=" * 80)
    print("Step 1: Navigate and search")
    print("=" * 80)
    facade.navigate_to_sales_order()
    facade.click_search()

    print("\n" + "=" * 80)
    print("Step 2: Get system order numbers from page")
    print("=" * 80)

    order_numbers = []
    max_retries = 3
    for attempt in range(max_retries):
        order_numbers = facade.get_sorted_order_numbers(5)
        if len(order_numbers) >= 1:
            break
        print(f"第 {attempt + 1} 次尝试获取订单数据，等待重试...")
        facade.order_page.wait_for_table_data()

    print(f"Page system order numbers: {len(order_numbers)} found")

    assert len(order_numbers) >= 1, f"Not enough orders found: {len(order_numbers)}"

    print("\n" + "=" * 80)
    print("Step 3: Select first 3 orders")
    print("=" * 80)

    selected_count = facade.order_page.select_first_orders(3)
    print(f"Selected {selected_count} orders")
    assert selected_count == min(
        3, len(order_numbers)
    ), f"实际选中{selected_count}条，期望选中{min(3, len(order_numbers))}条"

    print("\n" + "=" * 80)
    print("Step 4: Navigate to export page directly")
    print("=" * 80)

    order_param = ",".join(order_numbers[:3])
    # Resolve the URL from the active environment configuration.  Reading
    # TEST_WEB_BASE_URL here made a UAT run jump back to the test host when
    # both environments were present in the same .env file.
    base = get_config().base_url
    if base.endswith("/index"):
        base = base[: -len("/index")]
    base = base.rstrip("/")
    if not base:
        raise RuntimeError("当前环境 UI base_url 未设置，请检查 configs/uat.yaml 或对应环境变量")
    export_url = f"{base}/sales/order/exportPage?t={int(time.time()*1000)}&orderNo={order_param}"
    print("Export URL: same-origin export route reached")

    export_page = SalesOrderExportPage(page)
    page.goto(export_url)
    page.wait_for_load_state("networkidle")
    assert export_page.wait_for_export_page(timeout=30000), "Export page failed to load"

    print("\n" + "=" * 80)
    print("Step 5: Select template")
    print("=" * 80)
    result = export_page.select_export_template(EXPORT_TEMPLATE)
    print(f"Template selection result: {result}")
    assert result, f"模板选择失败: {result}"
    export_page.wait_for_page_settle(timeout=10000)

    print("\n" + "=" * 80)
    print("Step 6: Check current field selection and ensure 系统单号 is checked")
    print("=" * 80)

    field_status = page.evaluate("""
        () => {
            const labels = document.querySelectorAll('.el-checkbox__label');
            const result = [];
            for (let i = 0; i < labels.length; i++) {
                const text = labels[i].innerText.trim();
                if (text && text.length < 30) {
                    const cb = labels[i].closest('.el-checkbox')?.querySelector('input[type="checkbox"]');
                    result.push({
                        index: i,
                        text: text,
                        checked: cb ? cb.checked : false
                    });
                }
            }
            return result;
        }
    """)

    print(f"\nTotal fields: {len(field_status)}")
    print("\nFirst 15 fields:")
    for f in field_status[:15]:
        status = "✓" if f["checked"] else "✗"
        print(f"  [{status}] {f['index']:2d}: {f['text']}")

    # 查找系统单号字段
    sys_field = None
    for f in field_status:
        if "系统单号" in f["text"]:
            sys_field = f
            break

    if sys_field:
        print(f"\nFound '系统单号' field: index={sys_field['index']}, checked={sys_field['checked']}")
        if not sys_field["checked"]:
            print("Clicking to select '系统单号'...")
        assert export_page.ensure_field_selected("系统单号"), "系统单号字段未选中"
    else:
        pytest.fail("导出字段列表中未找到系统单号")

    print("\n" + "=" * 80)
    print("Step 7: Click realtime export and download")
    print("=" * 80)

    os.makedirs(".runtime/downloads/regression", exist_ok=True)
    timestamp = int(time.time())
    save_path = f".runtime/downloads/regression/sales_order_sysno_{timestamp}.xlsx"

    download_result = export_page.download_to(save_path, timeout=120000)
    print(f"Download result: success={download_result.get('success')}, size={download_result.get('file_size', 0)}")
    assert download_result["success"], f"下载失败: {download_result.get('error')}"
    assert os.path.exists(save_path), f"下载文件不存在: {save_path}"

    print("\n" + "=" * 80)
    print("Step 8: Verify exported file")
    print("=" * 80)

    import openpyxl

    wb = openpyxl.load_workbook(save_path)
    ws = wb.active

    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

    headers = [str(cell.value or "").strip() for cell in ws[1]]
    print(f"\nAll headers ({len(headers)}):")
    for i, h in enumerate(headers, 1):
        print(f"  Col {i:2d}: {h}")

    sys_col = next((i + 1 for i, h in enumerate(headers) if "系统单号" in h), None)
    print(f"\n系统单号 column: {'Col ' + str(sys_col) if sys_col else 'NOT FOUND'}")
    assert sys_col is not None, "导出文件中未找到系统单号列"

    export_sys_nos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[sys_col - 1]
        if val:
            val_str = str(val).strip()
            if val_str.endswith(".0"):
                val_str = val_str[:-2]
            export_sys_nos.append(val_str)
            print("  <redacted order number>")

    print(f"\nPage system order numbers: {len(order_numbers[:3])} checked")
    print(f"Export system order numbers: {len(export_sys_nos)} found")
    expected_set = set(order_numbers[:3])
    actual_set = set(export_sys_nos)
    assert (
        actual_set == expected_set
    ), f"导出系统单号集合不一致: expected={sorted(expected_set)}, actual={sorted(actual_set)}"
    wb.close()
    print("\n" + "=" * 80)
    print("Test completed")
    print("=" * 80)
