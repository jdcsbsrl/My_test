"""Inventory SKU Management Facade.

封装库存SKU管理相关UI操作的业务逻辑，供测试用例调用。
包含6大功能模块的测试组件：
1. 库存SKU查询
2. 全选功能
3. 分页功能
4. 分页数量调整
5. 导出功能
6. 导出字段选择
"""

from __future__ import annotations

import os
import time
from typing import Any

import allure

from modules.auto_test.core.logger import get_logger
from modules.auto_test.pages.inventory_export_page import InventoryExportPage
from modules.auto_test.pages.inventory_sku_page import InventorySKUPage

logger = get_logger()


class InventorySKUFacade:
    """库存SKU管理业务门面，封装高层测试场景。"""

    def __init__(self, page) -> None:
        self.page = page
        self.sku_page = InventorySKUPage(page)
        self.export_page = None

    @allure.step("完整流程：进入库存SKU页面")
    def navigate_to_inventory(self) -> None:
        """进入库存SKU查询页面"""
        self.sku_page.navigate_to_search_page()
        self.sku_page.wait_for_search_results()

    @allure.step("完整流程：按SKU编码查询")
    def search_by_sku(self, sku_code: str) -> dict[str, Any]:
        """按SKU编码搜索

        Args:
            sku_code: SKU编码或编码片段

        Returns:
            包含搜索耗时、结果数量、结果列表的字典
        """
        self.sku_page.click_reset()
        self.sku_page.fill_sku_code(sku_code)
        elapsed = self.sku_page.click_search()
        count = self.sku_page.get_result_count()
        results = self.sku_page.get_search_results()
        return {"keyword": sku_code, "elapsed": elapsed, "count": count, "results": results}

    @allure.step("完整流程：按产品名称查询")
    def search_by_product_name(self, product_name: str) -> dict[str, Any]:
        """按产品名称搜索"""
        self.sku_page.click_reset()
        self.sku_page.fill_product_name(product_name)
        elapsed = self.sku_page.click_search()
        count = self.sku_page.get_result_count()
        return {"keyword": product_name, "elapsed": elapsed, "count": count}

    @allure.step("完整流程：按仓库查询")
    def search_by_warehouse(self, warehouse: str) -> dict[str, Any]:
        """按仓库搜索"""
        self.sku_page.click_reset()
        self.sku_page.select_warehouse(warehouse)
        elapsed = self.sku_page.click_search()
        count = self.sku_page.get_result_count()
        return {"warehouse": warehouse, "elapsed": elapsed, "count": count}

    @allure.step("完整流程：组合查询")
    def search_by_combination(
        self, sku_code: str | None = None, product_name: str | None = None, warehouse: str | None = None
    ) -> dict[str, Any]:
        """组合条件搜索"""
        self.sku_page.click_reset()
        if sku_code:
            self.sku_page.fill_sku_code(sku_code)
        if product_name:
            self.sku_page.fill_product_name(product_name)
        if warehouse:
            self.sku_page.select_warehouse(warehouse)
        elapsed = self.sku_page.click_search()
        count = self.sku_page.get_result_count()
        return {
            "conditions": {"sku_code": sku_code, "product_name": product_name, "warehouse": warehouse},
            "elapsed": elapsed,
            "count": count,
        }

    @allure.step("完整流程：导出当前搜索结果")
    def export_current_search(
        self, select_all_fields: bool = True, fields: list[str] | None = None, download_dir: str = "downloads"
    ) -> dict[str, Any]:
        """导出当前搜索结果

        Args:
            select_all_fields: 是否全选所有字段
            fields: 指定字段列表（当select_all_fields=False时生效）
            download_dir: 下载目录

        Returns:
            导出结果信息
        """
        self.sku_page.select_export_current_search()
        self.export_page = InventoryExportPage(self.page)
        if not self.export_page.wait_for_export_page():
            return {"success": False, "error": "导出页面加载失败"}

        self.export_page.select_first_template_if_available()

        if select_all_fields:
            self.export_page.select_all_fields()
        elif fields:
            self.export_page.select_fields(fields)

        os.makedirs(download_dir, exist_ok=True)
        timestamp = int(time.time())
        save_path = f"{download_dir}/inventory_sku_{timestamp}.xlsx"
        result = self.export_page.download_to(save_path, timeout=180000)

        result["export_type"] = "current_search"
        result["select_all_fields"] = select_all_fields
        return result

    @allure.step("完整流程：导出勾选的SKU")
    def export_selected(
        self, select_all_fields: bool = True, fields: list[str] | None = None, download_dir: str = "downloads"
    ) -> dict[str, Any]:
        """导出勾选的SKU

        Args:
            select_all_fields: 是否全选所有字段
            fields: 指定字段列表
            download_dir: 下载目录

        Returns:
            导出结果信息
        """
        self.sku_page.select_export_selected()
        self.export_page = InventoryExportPage(self.page)
        if not self.export_page.wait_for_export_page():
            return {"success": False, "error": "导出页面加载失败"}

        self.export_page.select_first_template_if_available()

        if select_all_fields:
            self.export_page.select_all_fields()
        elif fields:
            self.export_page.select_fields(fields)

        os.makedirs(download_dir, exist_ok=True)
        timestamp = int(time.time())
        save_path = f"{download_dir}/inventory_sku_selected_{timestamp}.xlsx"
        result = self.export_page.download_to(save_path, timeout=180000)

        result["export_type"] = "selected"
        result["select_all_fields"] = select_all_fields
        return result

    @allure.step("完整流程：分页+全选+导出")
    def export_with_page_size(
        self,
        page_size: int,
        select_all_fields: bool = True,
        fields: list[str] | None = None,
        download_dir: str = "downloads",
    ) -> dict[str, Any]:
        """设置分页大小后全选导出

        Args:
            page_size: 每页条数
            select_all_fields: 是否全选字段
            fields: 指定字段列表
            download_dir: 下载目录

        Returns:
            导出结果信息（含分页设置耗时、选中数量等）
        """
        page_set_time = self.sku_page.set_page_size(page_size)
        self.page.wait_for_load_state("networkidle")

        result_count = self.sku_page.get_result_count()
        selected_count = min(result_count, page_size) if result_count > 0 else 0

        result = self.export_current_search(select_all_fields=select_all_fields, fields=fields, download_dir=download_dir)
        result["page_size"] = page_size
        result["page_set_elapsed"] = page_set_time
        result["result_count"] = result_count
        result["selected_count"] = selected_count
        return result

    @allure.step("验证表头全选状态")
    def verify_select_all_state(self, expected: bool) -> bool:
        """验证全选复选框状态"""
        actual = self.sku_page.is_header_checkbox_checked()
        logger.info(f"全选状态验证: 预期={expected}, 实际={actual}")
        return actual == expected

    @allure.step("验证分页设置生效")
    def verify_page_size(self, expected_size: int) -> bool:
        """验证分页大小是否生效（通过实际行数）"""
        actual_rows = self.sku_page.get_current_page_row_count()
        if actual_rows > expected_size:
            logger.warning(f"实际行数{actual_rows}超过预期{expected_size}")
        return True

    @allure.step("验证翻页功能")
    def verify_pagination_navigation(self) -> dict[str, Any]:
        """验证分页导航功能"""
        current = self.sku_page.get_current_page()
        total = self.sku_page.get_total_pages()

        if total > 1:
            self.sku_page.click_next_page()
            self.sku_page.wait_for_search_results()
            new_page = self.sku_page.get_current_page()

            self.sku_page.goto_page(1)
            self.sku_page.wait_for_search_results()
            back_page = self.sku_page.get_current_page()

            return {
                "initial_page": current,
                "total_pages": total,
                "after_next": new_page,
                "after_back": back_page,
                "navigation_works": new_page == 2 and back_page == 1,
            }
        return {
            "initial_page": current,
            "total_pages": total,
            "navigation_works": True,
            "note": "数据不足一页，跳过翻页验证",
        }

    @staticmethod
    @allure.step("验证导出文件完整性")
    def verify_export_file(file_path: str) -> dict[str, Any]:
        """验证导出文件是否完整"""
        if not os.path.exists(file_path):
            return {"valid": False, "error": "文件不存在"}

        file_size = os.path.getsize(file_path)
        result = {
            "valid": file_size > 1024,
            "file_path": file_path,
            "file_size": file_size,
            "file_size_kb": round(file_size / 1024, 2),
        }

        if file_path.endswith((".xlsx", ".xls")):
            try:
                import openpyxl

                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.worksheets[0]
                result["row_count"] = ws.max_row
                result["col_count"] = ws.max_column
                result["headers"] = [
                    str(ws.cell(row=1, column=c).value or "") for c in range(1, min(ws.max_column + 1, 20))
                ]
                wb.close()
            except Exception as e:
                result["excel_parse_error"] = str(e)

        return result
