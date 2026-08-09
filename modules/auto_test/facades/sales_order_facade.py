"""Sales Order Management Facade.

封装销售订单管理相关UI操作的业务逻辑，供测试用例调用。
包含核心功能模块：
1. 销售订单查询
2. 排序功能
3. 全选功能
4. 分页功能
5. 导出功能（导出勾选的订单）
6. 导出字段选择
"""

from __future__ import annotations

import os
import time
from typing import Any

import allure

from modules.auto_test.core.logger import get_logger
from modules.auto_test.pages.sales_order_export_page import SalesOrderExportPage
from modules.auto_test.pages.sales_order_page import SalesOrderPage

logger = get_logger()


class SalesOrderFacade:
    """销售订单管理业务门面，封装高层测试场景。"""

    def __init__(self, page) -> None:
        self.page = page
        self.order_page = SalesOrderPage(page)
        self.export_page = None

    @allure.step("完整流程：进入销售订单页面")
    def navigate_to_sales_order(self) -> None:
        """进入销售订单页面"""
        self.order_page.navigate_to("sales/order/saleOrder")
        self.order_page.wait_for_load_state()
        self.order_page.wait_for_table_data()

    @allure.step("完整流程：点击搜索按钮")
    def click_search(self) -> None:
        """点击搜索按钮获取订单列表"""
        self.page.evaluate(
            """
            () => {
                const buttons = document.querySelectorAll('button');
                for (const btn of buttons) {
                    if (btn.innerText.includes('搜索')) {
                        btn.click();
                        return btn.innerText;
                    }
                }
                return null;
            }
        """
        )
        self.order_page.wait_for_table_data()
        logger.info("搜索完成")

    @allure.step("完整流程：选择排序方式")
    def select_sort(self, column_name: str, is_ascending: bool = True) -> None:
        """选择指定排序列和排序方向"""
        self.order_page.select_sort_order(column_name, is_ascending)
        self.order_page.wait_for_sort_complete()
        logger.info(f"排序完成: {column_name}, 升序: {is_ascending}")

    @allure.step("完整流程：设置分页大小")
    def set_page_size(self, page_size: int) -> float:
        """设置分页大小"""
        start_time = time.time()
        try:
            self.page.evaluate(
                f"""() => {{
                const selects = document.querySelectorAll('.el-select');
                for (let i = 0; i < selects.length; i++) {{
                    const text = selects[i].innerText.trim();
                    if (text === '10' || text === '20' || text === '50') {{
                        selects[i].click();
                        setTimeout(() => {{
                            const options = document.querySelectorAll('.el-select-dropdown__item');
                            for (const opt of options) {{
                                if (opt.textContent.trim() === '{page_size}') {{
                                    opt.click();
                                    break;
                                }}
                            }}
                        }}, 500);
                        break;
                    }}
                }}
            }}"""
            )
            self.page.wait_for_load_state("networkidle")
            elapsed = time.time() - start_time
            logger.info(f"设置分页{page_size}/页，耗时{elapsed:.2f}秒")
            return elapsed
        except Exception as e:
            logger.warning(f"设置分页大小失败: {e}")
            return 0.0

    @allure.step("完整流程：全选当前页订单")
    def select_all_current_page(self) -> int:
        """全选当前页订单"""
        self.order_page.select_all_current_page()
        selected_count = self.order_page.verify_selected_count()
        logger.info(f"已选中 {selected_count} 个订单")
        return selected_count

    @allure.step("完整流程：获取选中订单数量")
    def get_selected_count(self) -> int:
        """获取选中订单数量"""
        return self.order_page.get_selected_count()

    @allure.step("完整流程：获取排序后的订单号列表")
    def get_sorted_order_numbers(self, limit: int = 50) -> list[str]:
        """获取排序后的订单号列表"""
        return self.order_page.get_sorted_order_numbers(limit)

    @allure.step("完整流程：导出勾选的订单")
    def export_selected(
        self,
        select_all_fields: bool = False,
        fields: list[str] | None = None,
        download_dir: str = "downloads",
        template_name: str = "Dayone标准模板 --计算账单",
        ensure_fields: list[str] | None = None,
    ) -> dict[str, Any]:
        """导出勾选的订单

        Args:
            select_all_fields: 是否全选所有字段
            fields: 指定字段列表（当select_all_fields=False时生效）
            download_dir: 下载目录
            template_name: 导出模板名称
            ensure_fields: 确保勾选的字段列表（在模板基础上追加勾选）

        Returns:
            导出结果信息
        """
        self.order_page.select_export_selected()
        self.page.wait_for_load_state("networkidle")

        self.export_page = SalesOrderExportPage(self.page)
        if not self.export_page.wait_for_export_page():
            return {"success": False, "error": "导出页面加载失败"}

        if not self.select_export_template(template_name):
            logger.warning(f"未找到模板: {template_name}")

        if select_all_fields:
            self.export_page.select_all_fields()
        elif fields:
            self.export_page.select_fields(fields)

        if ensure_fields:
            for field in ensure_fields:
                self.export_page.select_field(field)
                time.sleep(0.5)

        os.makedirs(download_dir, exist_ok=True)
        timestamp = int(time.time())
        save_path = f"{download_dir}/sales_order_selected_{timestamp}.xlsx"
        result = self.export_page.download_to(save_path, timeout=180000)

        result["export_type"] = "selected"
        result["select_all_fields"] = select_all_fields
        result["template_name"] = template_name
        result["ensure_fields"] = ensure_fields
        return result

    @allure.step("完整流程：选择导出模板")
    def select_export_template(self, template_name: str) -> bool:
        """选择导出模板"""
        return self.export_page.select_export_template(template_name)

    @allure.step("完整流程：分页+排序+全选+导出")
    def export_with_sort_and_page_size(
        self,
        sort_column: str,
        is_ascending: bool = True,
        page_size: int = 50,
        select_all_fields: bool = True,
        fields: list[str] | None = None,
        download_dir: str = "downloads",
        template_name: str = "Dayone标准模板 --计算账单",
    ) -> dict[str, Any]:
        """完整导出流程：排序→设置分页→全选→导出

        Args:
            sort_column: 排序列名称
            is_ascending: 是否升序
            page_size: 每页条数
            select_all_fields: 是否全选字段
            fields: 指定字段列表
            download_dir: 下载目录
            template_name: 导出模板名称

        Returns:
            导出结果信息（含分页设置、选中数量、排序信息等）
        """
        self.select_sort(sort_column, is_ascending)

        page_set_time = self.set_page_size(page_size)
        self.page.wait_for_load_state("networkidle")

        self.select_all_current_page()
        selected_count = self.order_page.verify_selected_count()

        page_order_numbers = self.get_sorted_order_numbers(page_size)

        result = self.export_selected(
            select_all_fields=select_all_fields, fields=fields, download_dir=download_dir, template_name=template_name
        )
        result["sort_column"] = sort_column
        result["is_ascending"] = is_ascending
        result["page_size"] = page_size
        result["page_set_elapsed"] = page_set_time
        result["selected_count"] = selected_count
        result["page_order_numbers"] = page_order_numbers

        return result

    @staticmethod
    def verify_export_file(file_path: str) -> dict[str, Any]:
        """验证导出文件是否完整"""
        if not os.path.exists(file_path):
            return {"valid": False, "error": "文件不存在"}

        file_size = os.path.getsize(file_path)
        result = {
            "valid": file_size > 1024,
            "file_path": file_path,
            "file_size": file_size,
            "file_size_kb": round(file_size / 1024, 2),
        }

        if file_path.endswith((".xlsx", ".xls")):
            try:
                import openpyxl

                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.worksheets[0]
                result["row_count"] = ws.max_row
                result["col_count"] = ws.max_column
                result["headers"] = [
                    str(ws.cell(row=1, column=c).value or "") for c in range(1, min(ws.max_column + 1, 20))
                ]
                wb.close()
            except Exception as e:
                result["excel_parse_error"] = str(e)

        return result

    @staticmethod
    def verify_export_order_consistency(
        file_path: str,
        expected_order_numbers: list[str],
        order_number_column_name: str = "系统单号",
        deduplicate: bool = True,
    ) -> dict[str, Any]:
        """验证导出文件中的订单号顺序与页面排序是否一致

        Args:
            file_path: 导出文件路径
            expected_order_numbers: 页面上排序后的订单号列表（系统单号）
            order_number_column_name: 订单号列名（默认"系统单号"）
            deduplicate: 是否对导出的订单号去重（一个订单有多个SKU行时需要去重）

        Returns:
            验证结果
        """
        if not os.path.exists(file_path):
            return {"success": False, "error": "文件不存在"}

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path)
            ws = wb.worksheets[0]

            headers = [str(cell.value or "").strip() for cell in ws[1]]

            order_col_index = None
            for i, header in enumerate(headers):
                if order_number_column_name in header:
                    order_col_index = i + 1
                    break

            if order_col_index is None:
                for i, header in enumerate(headers):
                    if "系统单号" in header:
                        order_col_index = i + 1
                        break

            if order_col_index is None:
                for i, header in enumerate(headers):
                    if "订单" in header and ("号" in header or "id" in header.lower() or "name" in header.lower()):
                        order_col_index = i + 1
                        break

            if order_col_index is None:
                order_col_index = 1

            export_order_numbers = []
            for row in ws.iter_rows(min_row=2):
                cell_value = row[order_col_index - 1].value
                if cell_value:
                    order_str = str(cell_value).strip()
                    if order_str.endswith(".0"):
                        order_str = order_str[:-2]
                    export_order_numbers.append(order_str)

            wb.close()

            if not expected_order_numbers:
                return {
                    "success": True,
                    "note": "无预期订单号，跳过顺序验证",
                    "export_count": len(export_order_numbers),
                    "order_column": (
                        headers[order_col_index - 1] if order_col_index <= len(headers) else f"column_{order_col_index}"
                    ),
                }

            if not export_order_numbers:
                return {
                    "success": False,
                    "error": "导出文件中无订单数据",
                    "expected_count": len(expected_order_numbers),
                    "order_column": (
                        headers[order_col_index - 1] if order_col_index <= len(headers) else f"column_{order_col_index}"
                    ),
                }

            export_unique = []
            if deduplicate:
                seen = set()
                for num in export_order_numbers:
                    if num not in seen:
                        seen.add(num)
                        export_unique.append(num)
            else:
                export_unique = export_order_numbers

            matching_count = 0
            mismatched_positions = []

            for i, expected in enumerate(expected_order_numbers):
                if i < len(export_unique):
                    actual = export_unique[i]
                    if expected == actual:
                        matching_count += 1
                    else:
                        mismatched_positions.append({"position": i + 1, "expected": expected, "actual": actual})

            return {
                "success": len(mismatched_positions) == 0,
                "expected_count": len(expected_order_numbers),
                "export_count": len(export_order_numbers),
                "export_unique_count": len(export_unique),
                "matching_count": matching_count,
                "mismatched_positions": mismatched_positions,
                "deduplicated": deduplicate,
                "headers": headers[:10],
                "order_column": (
                    headers[order_col_index - 1] if order_col_index <= len(headers) else f"column_{order_col_index}"
                ),
            }

        except Exception as e:
            return {"success": False, "error": str(e)}
